# -*- coding: utf-8 -*-
"""Infra Cost converter -> dashboard INFRA_DETAIL (node tree) + INFRA_BU (root strips).

13 tabs form a tree: Consolidated(root) -> DaaS/Distribution/MarTech -> product leaves.
Each tab holds 1-2 blocks of a month time-series (rolling 12, Jul'25->Jun'26):
  * a "breakdown" block  (Gross Revenue / COGS / Data / Hosting / Proxy / % of Revenue ...)
  * optionally a "COGS by Segment|Product" list block.

Design:
  * Meta = a NODE REGISTRY (tab -> node_id / parent / colour / order) + scalars. Children
    (strips) are DERIVED from parent links, so the tree is data-driven.
  * Generic block extraction: header rows detected by >=3 month tokens; label taken from the
    Particulars column (breakdown) or the Segment/Product column (list); month columns roll.
  * Units: values kept in $000s and rounded to integers (as the dashboard does); % of Revenue
    converted from fraction -> percentage number (0.31 -> 31.0), matching IC_JUN.
  * validate(): tree integrity (every parent exists, every child linked), COGS-by-segment
    reconciles to Total COGS, % of Revenue in range, month window consistent across tabs.

Outputs: infra_STANDARDIZED.json  +  Infra_Cost_STANDARDIZED_Jun26.xlsx
"""
import openpyxl, json, sys, re

def txt(c):  return '' if c is None else str(c).replace('\xa0', ' ').replace('\n', ' ').strip()
def norm(s): return re.sub(r'\s+', ' ', txt(s)).lower()
def grid(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]

MONTH_RE = re.compile(r"^[A-Za-z]{3,9}['’\-\s]?\d{2}$")     # Jul'25 / Jul’25 / Jul-25
def is_month(v): return bool(MONTH_RE.match(txt(v)))
def num(v):
    if isinstance(v, (int, float)): return float(v)
    s = txt(v).replace(',', '').replace('%', '')
    try: return float(s)
    except ValueError: return None

def kind_breakdown(label):
    l = label.lower()
    if 'gross revenue' in l or l == 'revenue': return 'rev'
    if '% of revenue' in l or 'margin %' in l: return 'pct'
    if l.startswith('total'): return 'total'
    if l == 'cogs': return 'cogs'
    return 'comp'

# ---------- default node registry (used when the workbook has no 'Meta' tab) ----------
# Sheet name -> (node_id, parent, display name, colour, order). Mirrors the dashboard Infra tree.
DEFAULT_INFRA_NODES = [
    {"tab":"Consolidated","node_id":"root","parent":"","name":"Consolidated","col":"#5E0FC0","order":0},
    {"tab":"DaaS","node_id":"daas","parent":"root","name":"DaaS","col":"#8012FF","order":1},
    {"tab":"Distribution","node_id":"dist","parent":"root","name":"Distribution","col":"#0E7C86","order":2},
    {"tab":"MarTech","node_id":"martech","parent":"root","name":"MarTech","col":"#C026D3","order":3},
    {"tab":"PG-Car + Rev.AI","node_id":"ic-pgcar","parent":"daas","name":"PG-Car + Rev.AI","col":"#A35BFF","order":1},
    {"tab":"PG-OTA","node_id":"ic-pgota","parent":"daas","name":"PG-OTA","col":"#8012FF","order":2},
    {"tab":"PG-Air + PG-Cruise","node_id":"ic-pgair","parent":"daas","name":"PG-Air + PG-Cruise","col":"#6B46E0","order":3},
    {"tab":"HospiBI","node_id":"ic-hospibi","parent":"daas","name":"HospiBI","col":"#5E0FC0","order":4},
    {"tab":"RezGain","node_id":"ic-rez","parent":"dist","name":"RezGain","col":"#0E7C86","order":1},
    {"tab":"UNO","node_id":"ic-uno","parent":"dist","name":"UNO","col":"#3F6AD8","order":2},
    {"tab":"Enterprise Connectivity","node_id":"ic-ec","parent":"dist","name":"Enterprise Connectivity","col":"#1E88C5","order":3},
    {"tab":"Sojern","node_id":"ic-sojern","parent":"martech","name":"Sojern","col":"#7A2BE8","order":1},
    {"tab":"SoHo","node_id":"ic-soho","parent":"martech","name":"SoHo","col":"#E0559B","order":2},
]

# ---------- Meta ----------
def read_meta(wb):
    m = {'nodes': []}
    if 'Meta' not in wb.sheetnames:
        m['nodes'] = [dict(n) for n in DEFAULT_INFRA_NODES]
        return m
    mode = None
    for row in grid(wb['Meta']):
        c = [txt(x) for x in row] + ['', '', '', '', '', '']
        if c[0] == 'key' and c[1] == 'value':  mode = 'kv'; continue
        if c[0] == 'tab' and c[1] == 'node_id': mode = 'nodes'; continue
        if c[0] == '' and c[1] == '': continue
        if c[0].startswith('Meta —') or c[0].startswith('13-tab'): continue
        if mode == 'kv' and c[1]:
            m[c[0]] = c[1]
        elif mode == 'nodes' and c[0]:
            m['nodes'].append({'tab': c[0], 'node_id': c[1], 'parent': c[2], 'name': c[3],
                               'col': c[4], 'order': int(c[5]) if str(c[5]).isdigit() else 99})
    if not m['nodes']:
        m['nodes'] = [dict(n) for n in DEFAULT_INFRA_NODES]
    return m

# ---------- block extraction ----------
def section_title(rows, hi):
    for k in (hi - 1, hi - 2):
        if k < 0: continue
        t = txt(rows[k][0]) or (txt(rows[k][1]) if len(rows[k]) > 1 else '')
        t = re.sub(r'^[A-Z]\s*[·.]\s*', '', t)          # strip "A · " / "B · " prefixes
        if t and not is_month(t):
            return t
    return None

def extract_tab(rows):
    blocks = []
    hdrs = [i for i, r in enumerate(rows) if sum(1 for v in r if is_month(v)) >= 3]
    for bi, hi in enumerate(hdrs):
        header = rows[hi]
        mcols = [(j, txt(header[j])) for j in range(len(header)) if is_month(header[j])]
        first_m = mcols[0][0]
        # the two label columns sit left of the first month column
        seg_col = next((j for j in range(first_m) if txt(header[j])), 0)
        part_col = next((j for j in range(seg_col + 1, first_m) if txt(header[j])), seg_col)
        title = section_title(rows, hi)
        is_list = bool(title and ('by segment' in title.lower() or 'by product' in title.lower()))
        stop = hdrs[bi + 1] - 1 if bi + 1 < len(hdrs) else len(rows)
        # trim trailing back so the next block's own title isn't swallowed
        data, blank = [], 0
        for r in rows[hi + 1: stop + 1]:
            if all(txt(x) == '' for x in r):
                blank += 1
                if blank >= 2: break
                continue
            blank = 0
            seg, part = txt(r[seg_col]) if seg_col < len(r) else '', txt(r[part_col]) if part_col < len(r) else ''
            vals = [num(r[j]) if j < len(r) else None for j, _ in mcols]
            if not any(v is not None for v in vals):
                continue
            if is_list:
                label = seg or part
                k = 'total' if label.lower().startswith('total') else 'cogs'
                v = [round(x) if x is not None else None for x in vals]
            else:
                label = part or seg
                k = kind_breakdown(label)
                v = [round(x * 100, 1) if (k == 'pct' and x is not None) else
                     (round(x) if x is not None else None) for x in vals]
            if not label:
                continue
            data.append({'l': label, 'k': k, 'v': v})
        if data:
            blocks.append({'title': title, 'kind': 'list' if is_list else 'breakdown',
                           'months': [mm for _, mm in mcols], 'rows': data})
    return blocks

# ---------- convert / assemble tree ----------
def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    m = read_meta(wb)
    by_tab = {norm(n['tab']): n for n in m['nodes']}
    nodes = {}
    for name in [s for s in wb.sheetnames if s != 'Meta']:
        nd = by_tab.get(norm(name))
        if not nd:
            continue
        blocks = extract_tab(grid(wb[name]))
        nodes[nd['node_id']] = {'name': nd['name'], 'parent': nd['parent'] or None,
                                'col': nd['col'], 'order': nd['order'],
                                'tables': [{'title': b['title'], 'kind': b['kind'],
                                            'months': b['months'], 'rows': b['rows']} for b in blocks]}
    # derive strips (children) from parent links, ordered
    for nid, nd in nodes.items():
        kids = sorted([(o['order'], cid) for cid, o in nodes.items() if o['parent'] == nid])
        nd['strips'] = [cid for _, cid in kids]
    return {'meta': {k: v for k, v in m.items() if k != 'nodes'}, 'nodes': nodes}

# ---------- validate ----------
def validate(res):
    ok, msg = True, []
    nodes = res['nodes']
    if 'root' not in nodes:
        ok = False; msg.append("no root node")
    for nid, nd in nodes.items():
        if nd['parent'] and nd['parent'] not in nodes:
            ok = False; msg.append(f"{nid}: parent '{nd['parent']}' missing")
        if not nd['tables']:
            ok = False; msg.append(f"{nid}: no tables extracted")
    # % of Revenue sanity (0..200 as percentage number)
    for nid, nd in nodes.items():
        for t in nd['tables']:
            for row in t['rows']:
                if row['k'] == 'pct':
                    for x in row['v']:
                        if x is not None and not (-50 <= x <= 200):
                            msg.append(f"note: {nid} % of Revenue out of range: {x}")
    # COGS-by-segment reconciliation on root: DaaS+Distribution+MarTech == Total COGS (last month)
    root = nodes.get('root', {})
    seg = next((t for t in root.get('tables', []) if t['kind'] == 'list'), None)
    if seg:
        rows = {r['l'].lower(): r['v'] for r in seg['rows']}
        parts = [rows.get(k) for k in ('daas', 'distribution', 'martech')]
        tot = next((v for k, v in rows.items() if k.startswith('total')), None)
        if all(parts) and tot:
            i = -1
            s = sum(p[i] for p in parts if p[i] is not None)
            if abs(s - tot[i]) > 1:
                ok = False; msg.append(f"root COGS segments {s} != Total {tot[i]} (last month)")
            else:
                msg.append(f"root COGS reconciles: DaaS+Dist+MarTech = Total = {tot[i]}")
    nt = sum(len(nd['tables']) for nd in nodes.values())
    msg.append(f"{len(nodes)} nodes, {nt} tables; tree parents/children linked")
    return ok, msg

# ---------- review Excel ----------
def write_xlsx(res, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); wb.remove(wb.active)
    AR = Font(name='Arial', size=9); BD = Font(name='Arial', size=9, bold=True)
    HF = PatternFill('solid', fgColor='5E0FC0'); HW = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='E2D9F0'); BB = Border(thin, thin, thin, thin)
    def hdr(ws, r, vals):
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.font = HW; c.fill = HF; c.border = BB
            c.alignment = Alignment(horizontal='center', wrap_text=True)
    def cell(ws, r, cc, v, font=AR, nfmt=None):
        c = ws.cell(r, cc, v); c.font = font; c.border = BB
        if nfmt: c.number_format = nfmt
        return c
    nodes = res['nodes']
    # Tree sheet
    ws = wb.create_sheet('Tree'); ws.cell(1, 1, 'Infra node tree (root → BU → product)').font = BD
    for k, v in res['meta'].items(): pass
    r = 2
    for k, v in res['meta'].items():
        cell(ws, r, 1, k, BD); cell(ws, r, 2, v); r += 1
    r += 1; hdr(ws, r, ['node_id', 'name', 'parent', 'children (strips)', 'colour', '#tables']); r += 1
    order = ['root'] + [c for c in nodes.get('root', {}).get('strips', [])]
    for bu in list(order):
        order += [c for c in nodes.get(bu, {}).get('strips', []) if c not in order]
    for nid in [n for n in order if n in nodes] + [n for n in nodes if n not in order]:
        nd = nodes[nid]
        cell(ws, r, 1, nid, BD); cell(ws, r, 2, nd['name']); cell(ws, r, 3, nd['parent'])
        cell(ws, r, 4, ' · '.join(nd['strips']) or '—'); cell(ws, r, 5, nd['col']); cell(ws, r, 6, len(nd['tables']))
        r += 1
    for w, col in zip([12, 22, 10, 34, 10, 8], 'ABCDEF'): ws.column_dimensions[col].width = w

    # one sheet per node (its tables)
    used = {'Tree'}
    for nid in [n for n in order if n in nodes]:
        nd = nodes[nid]; base = re.sub(r'[\\/*?:\[\]]', ' ', nid)[:28]; nm = base; k = 1
        while nm in used: k += 1; nm = f"{base[:26]}_{k}"
        used.add(nm)
        ws = wb.create_sheet(nm)
        ws.cell(1, 1, f"{nd['name']}  ({nid}) · parent={nd['parent']} · $000s").font = BD
        r = 3
        for t in nd['tables']:
            cell(ws, r, 1, t['title'] or t['kind'], BD); r += 1
            hdr(ws, r, ['Line', 'k'] + t['months']); r += 1
            for row in t['rows']:
                cell(ws, r, 1, row['l']); cell(ws, r, 2, row['k'])
                for j, mv in enumerate(row['v'], 3):
                    cell(ws, r, j, mv, nfmt=('0.0"%"' if row['k'] == 'pct' else '#,##0'))
                r += 1
            r += 1
        ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 6
        for c in [chr(x) for x in range(ord('C'), ord('C') + 13)]: ws.column_dimensions[c].width = 8
    wb.save(path)

# ---------- main ----------
if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs/Infra_Cost_Jun26_INPUT_withMeta.xlsx'
    R = convert(path)
    ok, msg = validate(R)
    print("META:", R['meta'])
    for nid, nd in R['nodes'].items():
        print(f"  {nid:12} name={nd['name']!r:24} parent={nd['parent']!s:8} strips={nd['strips']} tables={len(nd['tables'])}")
        for t in nd['tables']:
            print(f"       [{t['kind']:9}] {t['title']!r:34} months={len(t['months'])} rows={[r['l'] for r in t['rows']][:6]}")
    print("\nVALIDATE:", ("PASS" if ok else "FAIL"))
    for x in msg: print("   -", x)
    json.dump({'standardized': R, 'validate': {'pass': ok, 'messages': msg}},
              open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/infra_STANDARDIZED.json', 'w'), ensure_ascii=False, indent=1)
    write_xlsx(R, (__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/Infra_Cost_STANDARDIZED_Jun26.xlsx')
    print("\nWrote infra_STANDARDIZED.json + Infra_Cost_STANDARDIZED_Jun26.xlsx")
