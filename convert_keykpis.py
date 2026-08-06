# -*- coding: utf-8 -*-
"""Key KPIs converter. Single matrix: metrics as ROWS, BUs as COLUMNS.
Each metric has its own UNIT (percent/currency/count/multiple/months/ratio) declared in Meta
(with suffix-inference fallback). Parses '1.72x'->1.72, 'n/a'/'-'->null, normalizes nbsp."""
import openpyxl, json, sys, re

def txt(c): return '' if c is None else str(c).replace('\xa0',' ').strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_column and ws.max_column and ws.max_row+1)]
def norm(s): return re.sub(r'\s+',' ',txt(s)).lower()

def parse_val(c):
    if isinstance(c,(int,float)): return round(float(c),4)
    s=txt(c)
    if s.lower() in ('n/a','na','-','n/m',''): return None
    m=re.match(r'^-?\d[\d,]*\.?\d*\s*x$', s, re.I)     # "1.72x"
    if m: 
        try: return round(float(s.lower().replace('x','').replace(',','').strip()),4)
        except: return None
    try: return round(float(s.replace(',','')),4)
    except: return None

def infer_unit(label):
    l=norm(label)
    if 'count' in l or '(#)' in l: return 'count'
    if '($)' in l or 'per employee' in l or 'per client' in l: return 'currency'
    if '(x)' in l or 'multiple' in l: return 'multiple'
    if 'month' in l: return 'months'
    if 'ltv' in l or 'cac' in l: return 'ratio'
    if '(%)' in l or '% of rev' in l or 'rule' in l or 'rate' in l or 'monetization' in l: return 'percent'
    return 'number'

def read_meta(wb):
    m={'metrics':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='metric' and b=='unit': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['metrics'][norm(a)]=b
    return m

def unit_for(label, meta):
    n=norm(label)
    for k,u in meta['metrics'].items():
        if n.startswith(k) or k.startswith(n[:len(k)]) or k in n: return u
    return infer_unit(label)

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    ws=next(wb[t] for t in wb.sheetnames if t!='Meta'); rows=grid(ws)
    # header row = the one listing BU names (>=3 non-empty labels after col with blank)
    hi=next(i for i,r in enumerate(rows) if sum(1 for x in r if txt(x))>=4 and not any(isinstance(x,(int,float)) for x in r))
    H=[txt(x) for x in rows[hi]]
    bu_cols=[j for j,h in enumerate(H) if h and j>=2]
    bus=[H[j] for j in bu_cols]
    out=[]; section=None
    for row in rows[hi+1:]:
        label=txt(row[1]) if len(row)>1 else (txt(row[0]) if row else '')
        # some labels sit in col A or B
        if not label:
            label=next((txt(x) for x in row[:2] if txt(x)),'')
        if not label: continue
        vals=[parse_val(row[j]) if j<len(row) else None for j in bu_cols]
        if all(v is None for v in vals):
            section=label; continue                     # section header (e.g. '40% Rule Check')
        rec={'metric':label.lstrip('- ').strip(),'section':section,'unit':unit_for(label,m),
             'values':{bus[k]:vals[k] for k in range(len(bus))}}
        out.append(rec)
    return {'meta':{k:v for k,v in m.items() if k!='metrics'}, 'bus':bus, 'metrics':out}

def validate(res):
    ok=True; msg=[]
    if len(res['bus'])<3: ok=False; msg.append(f"only {len(res['bus'])} BUs found")
    for r in res['metrics']:
        if r['unit'] in ('number',): msg.append(f"unit not resolved: {r['metric']}")
        if all(v is None for v in r['values'].values()): ok=False; msg.append(f"{r['metric']}: all values null")
    return ok,(msg or ['all metrics parsed with a unit; BUs present'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/Key_KPIs_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'],"| BUs:",R['bus'])
    for r in R['metrics']: print(f"  [{r['unit']:8}] {(r['section']+' / ') if r['section'] else ''}{r['metric']}: {list(r['values'].values())}")
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:6]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/keykpis_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
