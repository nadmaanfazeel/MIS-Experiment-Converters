# -*- coding: utf-8 -*-
"""Executive Summary converter  (NARRATIVE / commentary folder).

Unlike the numeric folders, this is prose. The converter does three things:
  1. FAITHFUL CAPTURE of the section -> bullet (-> sub-bullet, on the Sojern tab)
     hierarchy, verbatim text preserved. Sections are the bold labels in col A;
     bullets carry a '*'/• marker (level 1 in col A, level 2 '○' in col B on Sojern).
  2. MENTIONS layer: pulls $ / % figures out of the prose with light context
     (metric, YTD/YoY/vs-budget) -> a queryable table for the future text2SQL bot.
     These are TEXT-EXTRACTED and NON-AUTHORITATIVE (authoritative numbers come
     from the numeric converters, e.g. P&L); flagged as such.
  3. ROUTING into the dashboard ES payload shape (consolidated{stats,narr} +
     bus{daas,dist,martech}{headline,teaser,products[{name,metric,points}]}),
     driven entirely by the Meta section-map + product-map. Final prose polish
     (bolding entities, light rewording) is the human/LLM review step.

Outputs: exec_summary_STANDARDIZED.json  +  Executive_Summary_STANDARDIZED_Jun26.xlsx
"""
import openpyxl, json, sys, re

BULLETS = {'•', '·', '-', '*', 'o', '○', '◦', '▪'}
def txt(c):  return '' if c is None else str(c).replace('\xa0', ' ').strip()
def norm(s): return re.sub(r'[^a-z0-9]', '', txt(s).lower())        # aggressive key
def nsp(s):  return re.sub(r'\s+', ' ', txt(s)).strip()
def grid(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]

SCALE = {'k': 1e3, 'm': 1e6, 'mn': 1e6, 'bn': 1e9}

# ---------- Meta ----------
def read_meta(wb):
    m = {'sections': [], 'products': []}
    if 'Meta' not in wb.sheetnames:
        return m
    mode = None
    for row in grid(wb['Meta']):
        c = [txt(x) for x in row] + ['', '', '', '', '']
        a = c[0]
        if a == 'key' and c[1] == 'value':       mode = 'kv'; continue
        if a == 'section' and c[1] == 'tab':      mode = 'sec'; continue
        if a == 'label' and c[1] == 'bu':         mode = 'prod'; continue
        if a == '' and c[1] == '':                continue
        if a.lower().startswith(('meta —', 'this folder')): continue
        if mode == 'kv' and c[1]:
            m[a] = c[1]
        elif mode == 'sec' and a:
            m['sections'].append({'section': a, 'tab': c[1], 'bu': c[2], 'role': c[3], 'product': c[4]})
        elif mode == 'prod' and a:
            m['products'].append({'label': a, 'bu': c[1], 'dash_product': c[2],
                                  'order': int(c[3]) if str(c[3]).isdigit() else 99})
    return m

# ---------- mention extraction ----------
def extract_mentions(text):
    out = []
    low = text.lower()
    for mm in re.finditer(r'\(?\$?\s*(-?[\d,]+(?:\.\d+)?)\s*(k|mn|m|bn)?\)?\s*(%)?', text):
        num, sc, pct = mm.group(1), (mm.group(2) or '').lower(), mm.group(3)
        raw = mm.group(0).strip()
        if not re.search(r'[\d]', num):
            continue
        if not (('$' in raw) or pct or sc):        # skip bare integers like years / counts w/o unit cue
            if not re.search(r'\baccounts?\b|\badds?\b|churn', low):   # keep count-ish only w/ cue
                continue
        val = float(num.replace(',', ''))
        if pct:
            unit, scale, vnorm = 'percent', None, val / 100.0
        elif '$' in raw or sc:
            unit, scale = 'currency', (sc or None)
            vnorm = val * SCALE.get(sc, 1)
        else:
            unit, scale, vnorm = 'count', None, val
        # context: basis from a symmetric window, metric from the PRECEDING words
        st = max(0, mm.start() - 40); win = low[st:mm.end() + 40]
        pre = low[max(0, mm.start() - 30):mm.start()]
        basis = []
        if 'budget' in win or 'plan' in win or 'v/s' in win or ' vs' in win or 'against' in win: basis.append('vs_budget')
        if 'ytd' in win: basis.append('ytd')
        if 'yoy' in win or 'y-o-y' in win or 'last year' in win: basis.append('yoy')
        metric = ('ebitda' if ('ebitda' in pre or 'fcf' in pre or 'free cash flow' in pre)
                  else 'revenue' if ('revenue' in pre or 'gross rev' in pre or re.search(r'\brev\b', pre))
                  else 'expenses' if ('expense' in pre or 'cost' in pre) else None)
        out.append({'raw': raw, 'value': val, 'unit': unit, 'scale': scale,
                    'value_norm': vnorm, 'metric': metric, 'basis': basis})
    return out

# ---------- parse a tab into sections -> bullets ----------
def is_section(a_val, a_font_bold, has_bullet):
    a = txt(a_val)
    return bool(a) and a_font_bold and not has_bullet and a not in BULLETS

def parse_tab(wsv, wsf, meta_sections):
    rowsv = grid(wsv)
    sec_by_norm = {norm(s['section']): s for s in meta_sections}
    title = ''
    sections, cur = [], None
    for i, row in enumerate(rowsv):
        a = txt(row[0]); b = txt(row[1]) if len(row) > 1 else ''; cc = txt(row[2]) if len(row) > 2 else ''
        bold = bool(wsf.cell(i + 1, 1).font and wsf.cell(i + 1, 1).font.bold)
        if not a and not b and not cc:
            continue
        # title = the first big bold header (size >=14 or matches a Sojern/ES title)
        f1 = wsf.cell(i + 1, 1).font
        if not title and bold and f1 and (f1.size or 0) >= 12:
            title = a; continue
        marker_A = a in BULLETS
        marker_B = b in BULLETS
        if is_section(a, bold, marker_A):
            sdef = sec_by_norm.get(norm(a), {'bu': '', 'role': '', 'product': ''})
            cur = {'section': a, **sdef, 'bullets': []}
            sections.append(cur); continue
        # bullet rows
        if marker_A:                                   # level-1 bullet, text in col B (or C)
            body = b or cc
            if cur is None:                            # bullet before any section header
                sdef = sec_by_norm.get(norm(title), {'bu': '', 'role': '', 'product': ''})
                cur = {'section': title or 'General', **sdef, 'bullets': []}
                sections.append(cur)
            cur['bullets'].append({'level': 1, 'text': body})
        elif marker_B:                                 # level-2 sub-bullet (Sojern), text in col C
            if cur and cur['bullets']:
                cur['bullets'][-1].setdefault('sub', []).append(cc)
            elif cur:
                cur['bullets'].append({'level': 2, 'text': cc})
    return title, sections

# ---------- sublabel + routing ----------
def split_sublabel(text, prod_by_norm):
    """If the bullet starts with 'Label:' and Label matches a product-map entry, split it."""
    m = re.match(r'\s*([^:\n]{2,40}?):\s*(.*)', text, re.S)
    if m:
        lead = m.group(1).strip()
        if norm(lead) in prod_by_norm:
            return prod_by_norm[norm(lead)], m.group(2).strip()
    return None, text

def convert(path):
    wbv = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path, data_only=False)
    m = read_meta(wbv)
    prod_by_norm = {norm(p['label']): p for p in m['products']}
    tabs = []
    for name in [t for t in wbv.sheetnames if t != 'Meta']:
        title, sections = parse_tab(wbv[name], wbf[name], m['sections'])
        # attach mentions + sublabels
        for s in sections:
            for blt in s['bullets']:
                prod, rest = split_sublabel(blt['text'], prod_by_norm)
                blt['sublabel'] = prod['label'] if prod else None
                blt['dash_product'] = prod['dash_product'] if prod else None
                blt['mentions'] = extract_mentions(blt['text'])
                for sub in blt.get('sub', []):
                    blt['mentions'] += extract_mentions(sub)
        tabs.append({'tab': name, 'title': title, 'sections': sections})
    return {'meta': {k: v for k, v in m.items() if k not in ('sections', 'products')},
            'section_map': m['sections'], 'product_map': m['products'], 'tabs': tabs}

# ---------- assemble dashboard ES preview (Meta-driven) ----------
def clean(t):  return nsp(t.replace('\n', ' '))

def assemble_es(res):
    prod_by_norm = {norm(p['label']): p for p in res['product_map']}
    ES = {'consolidated': {'stats': [], 'narr': []}, 'bus': {}}
    order = res['meta'].get('bu_order', 'daas,dist,martech').split(',')
    for bu in order:
        ES['bus'][bu] = {'name': bu, 'headline': '', 'teaser': '', 'products': []}
    def get_product(bu, dash_name, order_hint=99):
        for p in ES['bus'][bu]['products']:
            if p['name'] == dash_name:
                return p
        p = {'name': dash_name, 'order': order_hint, 'metric': '', 'points': []}
        ES['bus'][bu]['products'].append(p); return p

    for tab in res['tabs']:
        for s in tab['sections']:
            bu, role = s.get('bu', ''), (s.get('role') or '')
            roles = role.split('+')
            cur_prod = None
            for blt in s['bullets']:
                body = clean(blt['text'])
                sub = [clean(x) for x in blt.get('sub', [])]
                if 'narrative' in roles:                       # Consolidated
                    ES['consolidated']['narr'].append(body)
                    continue
                # a product sub-label starts / selects a product
                if blt.get('dash_product'):
                    pm = prod_by_norm.get(norm(blt['sublabel']))
                    cur_prod = get_product(bu, blt['dash_product'], pm['order'] if pm else 99)
                    cur_prod['points'].append(clean(re.sub(r'^\s*[^:\n]{2,40}?:\s*', '', blt['text'], flags=re.S)))
                    cur_prod['points'] += sub
                elif 'product' in roles and s.get('product'):   # whole section is one product (Hospi BI, SoHo, Sojern)
                    pm = prod_by_norm.get(norm(s['product']))
                    cur_prod = get_product(bu, s['product'], pm['order'] if pm else 99)
                    cur_prod['points'].append(body); cur_prod['points'] += sub
                elif 'headline' in roles and cur_prod is None and not ES['bus'][bu]['headline']:
                    ES['bus'][bu]['headline'] = body            # first unlabeled bullet = BU headline
                elif cur_prod is not None:
                    cur_prod['points'].append(body); cur_prod['points'] += sub
                else:
                    ES['bus'][bu]['headline'] = (ES['bus'][bu]['headline'] + ' ' + body).strip()
    # sort products by order; derive a teaser from headline
    for bu in ES['bus'].values():
        bu['products'].sort(key=lambda p: p['order'])
        bu['teaser'] = (bu['headline'][:120] + '…') if len(bu['headline']) > 120 else bu['headline']
    # consolidated stat cards (best-effort from Consolidated + Sojern mentions)
    ES['consolidated']['stats'] = _consol_stats(res)
    return ES

def _consol_stats(res):
    def find_section(pred):
        for tab in res['tabs']:
            for s in tab['sections']:
                if pred(s): return s
        return None
    stats = []
    mon = res['meta'].get('month', '')
    cons = find_section(lambda s: (s.get('role') or '') == 'narrative')
    if cons and cons['bullets']:
        men = cons['bullets'][0]['mentions']
        rev = next((x for x in men if x['unit'] == 'currency' and x['metric'] == 'revenue'), None)
        eb  = next((x for x in men if x['unit'] == 'currency' and x['metric'] == 'ebitda'), None)
        if rev: stats.append({'k': f'Group revenue · {mon}', 'v': f"${rev['value']}{rev['scale'] or ''}", 's': 'gross revenue'})
        if eb:  stats.append({'k': f'Group EBITDA · {mon}',  'v': f"${eb['value']}{eb['scale'] or ''}", 's': 'vs budget'})
    soj = find_section(lambda s: norm(s.get('product', '')) == 'sojern')
    if soj and soj['bullets']:
        men = soj['bullets'][0]['mentions']
        r = next((x for x in men if x['unit'] == 'currency'), None)
        if r: stats.append({'k': f'Sojern revenue · {mon}', 'v': f"${r['value']}{r['scale'] or ''}", 's': 'vs plan'})
    return stats

# ---------- validate ----------
def validate(res, ES):
    ok, msg = True, []
    declared = {norm(s['section']) for s in res['section_map']}
    found = {norm(s['section']) for tab in res['tabs'] for s in tab['sections']}
    missing = declared - found
    if missing:
        ok = False; msg.append(f"section(s) not found: {sorted(missing)}")
    # every section has at least one bullet
    for tab in res['tabs']:
        for s in tab['sections']:
            if not s['bullets']:
                ok = False; msg.append(f"{tab['tab']}/{s['section']}: no bullets")
    # every bullet routed to a BU
    orphan = [f"{t['tab']}/{s['section']}" for t in res['tabs'] for s in t['sections']
              if s['bullets'] and not s.get('bu')]
    if orphan:
        ok = False; msg.append(f"unrouted section(s): {sorted(set(orphan))}")
    # each mapped product should surface
    got = {p['name'] for bu in ES['bus'].values() for p in bu['products']}
    want = {p['dash_product'] for p in res['product_map']}
    miss_p = want - got
    if miss_p:
        msg.append(f"note: product(s) not populated: {sorted(miss_p)}")
    # Sojern 2-level parse present
    soj = next((t for t in res['tabs'] if t['tab'] == 'Sojern'), None)
    if soj and not any(b.get('sub') for s in soj['sections'] for b in s['bullets']):
        msg.append("note: no level-2 sub-bullets parsed on Sojern tab")
    empty_hl = [bu for bu, d in ES['bus'].items() if not d['headline']]
    if empty_hl:
        msg.append(f"note: no BU-level headline in source for {empty_hl} (write during review)")
    nb = sum(len(s['bullets']) for t in res['tabs'] for s in t['sections'])
    nm = sum(len(b['mentions']) for t in res['tabs'] for s in t['sections'] for b in s['bullets'])
    msg.append(f"{nb} bullets across {sum(len(t['sections']) for t in res['tabs'])} sections; {nm} figures extracted")
    return ok, msg

# ---------- review Excel ----------
def write_xlsx(res, ES, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); wb.remove(wb.active)
    AR = Font(name='Arial', size=10); BD = Font(name='Arial', size=10, bold=True)
    HF = PatternFill('solid', fgColor='1F4E78'); HW = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9D9D9'); BB = Border(thin, thin, thin, thin)
    TOP = Alignment(wrap_text=True, vertical='top')
    def hdr(ws, r, vals):
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.font = HW; c.fill = HF; c.border = BB
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    def cell(ws, r, cc, v, font=AR, wrap=False):
        c = ws.cell(r, cc, v); c.font = font; c.border = BB
        if wrap: c.alignment = TOP
        return c

    # Meta echo
    ws = wb.create_sheet('Meta'); ws.cell(1, 1, 'Meta').font = BD; r = 2
    for k, v in res['meta'].items():
        cell(ws, r, 1, k, BD); cell(ws, r, 2, v); r += 1
    r += 1; hdr(ws, r, ['section', 'tab', 'bu', 'role', 'product']); r += 1
    for s in res['section_map']:
        for j, key in enumerate(['section', 'tab', 'bu', 'role', 'product'], 1): cell(ws, r, j, s[key])
        r += 1
    r += 1; hdr(ws, r, ['label', 'bu', 'dash_product', 'order']); r += 1
    for p in res['product_map']:
        for j, key in enumerate(['label', 'bu', 'dash_product', 'order'], 1): cell(ws, r, j, p[key])
        r += 1
    for w, col in zip([26, 20, 12, 18, 22], 'ABCDE'): ws.column_dimensions[col].width = w

    # Narrative (faithful hierarchy)
    ws = wb.create_sheet('Narrative')
    hdr(ws, 1, ['Tab', 'Section', 'BU', 'Role', 'Lvl', 'Sub-label', 'Dash Product', '# Fig', 'Text (verbatim)']); r = 2
    for tab in res['tabs']:
        for s in tab['sections']:
            for b in s['bullets']:
                cell(ws, r, 1, tab['tab']); cell(ws, r, 2, s['section']); cell(ws, r, 3, s.get('bu'))
                cell(ws, r, 4, s.get('role')); cell(ws, r, 5, b['level']); cell(ws, r, 6, b.get('sublabel'))
                cell(ws, r, 7, b.get('dash_product')); cell(ws, r, 8, len(b['mentions']))
                joined = b['text'] + (('\n  ○ ' + '\n  ○ '.join(b['sub'])) if b.get('sub') else '')
                cell(ws, r, 9, joined, wrap=True); r += 1
    for w, col in zip([18, 16, 9, 16, 5, 16, 20, 6, 90], 'ABCDEFGHI'): ws.column_dimensions[col].width = w

    # Mentions (queryable figures)
    ws = wb.create_sheet('Mentions')
    ws.cell(1, 1, 'TEXT-EXTRACTED figures — non-authoritative (authoritative numbers come from the numeric converters, e.g. P&L). For the future text2SQL layer & commentary sanity-checks.').font = AR
    hdr(ws, 3, ['Tab', 'Section', 'Sub-label', 'Metric', 'Basis', 'Value', 'Unit', 'Scale', 'Value (norm USD/frac)', 'Raw']); r = 4
    for tab in res['tabs']:
        for s in tab['sections']:
            for b in s['bullets']:
                for x in b['mentions']:
                    cell(ws, r, 1, tab['tab']); cell(ws, r, 2, s['section']); cell(ws, r, 3, b.get('sublabel'))
                    cell(ws, r, 4, x['metric']); cell(ws, r, 5, ', '.join(x['basis'])); cell(ws, r, 6, x['value'])
                    cell(ws, r, 7, x['unit']); cell(ws, r, 8, x['scale'] or '')
                    c = cell(ws, r, 9, x['value_norm']); c.number_format = '0.##%' if x['unit'] == 'percent' else '#,##0'
                    cell(ws, r, 10, x['raw']); r += 1
    for w, col in zip([16, 15, 16, 10, 16, 10, 9, 7, 18, 12], 'ABCDEFGHIJ'): ws.column_dimensions[col].width = w

    # ES_Preview (routed dashboard payload — DRAFT)
    ws = wb.create_sheet('ES_Preview')
    ws.cell(1, 1, 'Routed into the dashboard ES shape (DRAFT). Final bolding/rewording = the human/LLM review step.').font = AR
    r = 3
    cell(ws, r, 1, 'CONSOLIDATED — stat cards', BD); r += 1
    hdr(ws, r, ['k', 'v', 's']); r += 1
    for st in ES['consolidated']['stats']:
        cell(ws, r, 1, st['k']); cell(ws, r, 2, st['v']); cell(ws, r, 3, st['s'], wrap=True); r += 1
    r += 1; cell(ws, r, 1, 'CONSOLIDATED — narrative', BD); r += 1
    for n in ES['consolidated']['narr']:
        cell(ws, r, 1, n, wrap=True); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    r += 1
    for bu_key in res['meta'].get('bu_order', 'daas,dist,martech').split(','):
        bu = ES['bus'][bu_key]
        cell(ws, r, 1, f"BU: {bu_key.upper()}", BD); cell(ws, r, 2, bu['headline'], wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6); r += 1
        hdr(ws, r, ['#', 'Product', 'Points (joined)', '', '', '']); r += 1
        for p in bu['products']:
            cell(ws, r, 1, p['order']); cell(ws, r, 2, p['name'], BD)
            cell(ws, r, 3, '\n• ' + '\n• '.join(p['points']), wrap=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6); r += 1
        r += 1
    for w, col in zip([22, 26, 30, 12, 12, 12], 'ABCDEF'): ws.column_dimensions[col].width = w
    ws.column_dimensions['C'].width = 60
    wb.save(path)

# ---------- main ----------
if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs/Executive_Summary_Jun26_INPUT_withMeta.xlsx'
    R = convert(path)
    ES = assemble_es(R)
    ok, msg = validate(R, ES)
    print("META:", R['meta'])
    for tab in R['tabs']:
        print(f"\n=== TAB {tab['tab']} (title={tab['title']!r}) ===")
        for s in tab['sections']:
            print(f"  [{s.get('bu'):10}/{s.get('role'):16}] SECTION {s['section']!r}  ({len(s['bullets'])} bullets)")
            for b in s['bullets']:
                sl = f" sub={b['sublabel']}" if b.get('sublabel') else ""
                print(f"      L{b['level']}{sl}  figs={len(b['mentions'])}  {clean(b['text'])[:80]!r}")
    print("\n=== ES PREVIEW ===")
    print("consolidated.stats:", ES['consolidated']['stats'])
    for bu, d in ES['bus'].items():
        print(f"  {bu}: headline={d['headline'][:70]!r}")
        for p in d['products']:
            print(f"      - {p['name']}  ({len(p['points'])} pts)")
    print("\nVALIDATE:", ("PASS" if ok else "FAIL"), "|", '; '.join(msg))
    json.dump({'standardized': R, 'es_preview': ES, 'validate': {'pass': ok, 'messages': msg}},
              open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/exec_summary_STANDARDIZED.json', 'w'), ensure_ascii=False, indent=1)
    write_xlsx(R, ES, (__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/Executive_Summary_STANDARDIZED_Jun26.xlsx')
    print("\nWrote exec_summary_STANDARDIZED.json + Executive_Summary_STANDARDIZED_Jun26.xlsx")
