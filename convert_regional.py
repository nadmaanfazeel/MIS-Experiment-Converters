# -*- coding: utf-8 -*-
"""Regional Revenue converter. One sheet, 5 side-by-side region blocks
(Consol, NORAM, EU, APMEA, LATAM); each block = Product/BU | FY26 | YTD FY26 | YTD FY27 | Growth.
Hierarchy (bu/sub_bu/leaf/total) declared in Meta (indentation is inconsistent).
Values $'000; Growth fraction ('n/m' -> null)."""
import openpyxl, json, sys

def num(c):
    if isinstance(c,(int,float)): return round(float(c),3)
    return None
def txt(c): return '' if c is None else str(c).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

def read_meta(wb):
    m={'nodes':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='node' and b=='role': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['nodes'][a]=b
    return m

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    ws=next(wb[t] for t in wb.sheetnames if t!='Meta'); rows=grid(ws)
    # header row = the one with multiple 'Product/BU' cells
    hi=next(i for i,r in enumerate(rows) if sum('Product/BU' in txt(x) for x in r)>=2)
    label_cols=[j for j,x in enumerate(rows[hi]) if 'Product/BU' in txt(x)]
    # region name from the row above each label col
    regrow=rows[hi-1]
    regions=[]
    for lc in label_cols:
        nm=txt(regrow[lc]) if lc<len(regrow) else ''
        if 'Consol' in nm or nm=='': nm='Consol' if ('Consol' in nm or lc==label_cols[0]) else nm
        regions.append(nm)
    out=[]; cur_bu=None; cur_sub=None
    base=label_cols[0]
    for row in rows[hi+1:]:
        node=txt(row[base]) if base<len(row) else ''
        if not node: continue
        role=m['nodes'].get(node,'leaf')
        if role=='bu': cur_bu=node; cur_sub=None; level=0; parent=None
        elif role=='sub_bu': cur_sub=node; level=1; parent=cur_bu
        elif role=='total': level=0; parent=None; cur_bu=None; cur_sub=None
        else: level=2 if cur_sub else 1; parent=cur_sub or cur_bu; role='leaf'
        rec={'node':node,'role':role,'level':level,'parent':parent,'regions':{}}
        for reg,lc in zip(regions,label_cols):
            rec['regions'][reg]={'fy26':num(row[lc+1]) if lc+1<len(row) else None,
                                 'ytd_pfy':num(row[lc+2]) if lc+2<len(row) else None,
                                 'ytd_cur':num(row[lc+3]) if lc+3<len(row) else None,
                                 'growth':num(row[lc+4]) if lc+4<len(row) else None}
        out.append(rec)
    return {'meta':{k:v for k,v in m.items() if k!='nodes'}, 'regions':regions, 'rows':out}

def validate(res):
    ok=True; msg=[]; rows=res['rows']
    sub_regions=[r for r in res['regions'] if r!='Consol']
    # 1) Consol == sum of 4 sub-regions (ytd_cur) per node
    for r in rows:
        cons=r['regions'].get('Consol',{}).get('ytd_cur')
        s=sum((r['regions'].get(x,{}).get('ytd_cur') or 0) for x in sub_regions)
        if cons is not None and abs(cons-s)>max(1.0,0.01*abs(cons)):
            ok=False; msg.append(f"{r['node']}: Consol {round(cons,1)} != sum(regions) {round(s,1)}")
    # 2) within each region, bu/sub_bu == sum(children) (ytd_cur)
    def kids(p): return [x for x in rows if x['parent']==p]
    for r in rows:
        if r['role'] in ('bu','sub_bu'):
            ch=kids(r['node'])
            if ch:
                for reg in res['regions']:
                    pv=r['regions'][reg]['ytd_cur']; s=sum((c['regions'][reg]['ytd_cur'] or 0) for c in ch)
                    if pv is not None and abs(pv-s)>max(1.0,0.02*abs(pv)):
                        ok=False; msg.append(f"{reg}/{r['node']}: {round(pv,1)} != childsum {round(s,1)}"); break
    return ok,(msg or ['Consol=sum(regions) + hierarchy sums hold'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/Regional_Revenue_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'],"| regions:",R['regions'])
    print("rows:",len(R['rows']),"| bus:",[r['node'] for r in R['rows'] if r['role']=='bu'],"| subs:",[r['node'] for r in R['rows'] if r['role']=='sub_bu'])
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:6]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/regional_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
