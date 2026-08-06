# -*- coding: utf-8 -*-
"""Top Accounts converter (8 tabs, 2 schemas, MIXED currency).
- schema 'rg'     : #|Account|FY25-26 Act|YTD FY25-26|YTD FY26-27|YTD Bud|YoY%|Bud vs Act%|Region|Comments  ($'000)
- schema 'sojern' : #|Account|YTD FY26-27|YTD FY25-26|YTD FY24-25|YoY%(2027)|YoY%(2026)  (full $); may have 2027/2026 sub-tables
Meta supplies per-tab schema/bu/currency. 'N/m','-' -> null."""
import openpyxl, json, sys, re

def num(c):
    if isinstance(c,(int,float)): return round(float(c),3)
    return None
def txt(c): return '' if c is None else str(c).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

def read_meta(wb):
    m={'tabs':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='tab' and b=='schema': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['tabs'][a]={'schema':b,'bu':txt(row[2]) if len(row)>2 else '','currency':txt(row[3]) if len(row)>3 else ''}
    return m

def find_col(H,*preds):
    for pred in preds:
        for j,h in enumerate(H):
            if pred(h.lower()): return j
    return None

def parse_rg(rows,hi):
    H=[txt(x) for x in rows[hi]]
    c_rank=find_col(H,lambda h:h in ('#','s no.','s no','sno'))
    c_acc=find_col(H,lambda h:'primary account' in h)
    c_fy=find_col(H,lambda h:'fy 25-26' in h and 'ytd' not in h)
    c_ypfy=find_col(H,lambda h:'ytd fy 25-26' in h)
    c_ycur=find_col(H,lambda h:'ytd fy 26-27' in h and 'bud' not in h)
    c_ybud=find_col(H,lambda h:'ytd fy 26-27' in h and 'bud' in h)
    c_yoy=find_col(H,lambda h:'yoy' in h)
    c_bva=find_col(H,lambda h:'bud vs act' in h or 'bud vs' in h)
    c_reg=find_col(H,lambda h:'region' in h)
    c_com=find_col(H,lambda h:'comment' in h)
    acc=[]; summ=[]
    for row in rows[hi+1:]:
        rank=row[c_rank] if c_rank is not None and c_rank<len(row) else None
        label=txt(row[c_acc]) if c_acc is not None and c_acc<len(row) else ''
        if not label: continue
        rec={'account':label,
             'fy_pfy_act':num(row[c_fy]) if c_fy is not None and c_fy<len(row) else None,
             'ytd_pfy':num(row[c_ypfy]) if c_ypfy is not None and c_ypfy<len(row) else None,
             'ytd_cur':num(row[c_ycur]) if c_ycur is not None and c_ycur<len(row) else None,
             'ytd_bud':num(row[c_ybud]) if c_ybud is not None and c_ybud<len(row) else None,
             'yoy_pct':num(row[c_yoy]) if c_yoy is not None and c_yoy<len(row) else None,
             'bud_var_pct':num(row[c_bva]) if c_bva is not None and c_bva<len(row) else None,
             'region':txt(row[c_reg]) if c_reg is not None and c_reg<len(row) else '',
             'comments':txt(row[c_com]) if c_com is not None and c_com<len(row) else ''}
        if isinstance(rank,(int,float)): rec['rank']=int(rank); acc.append(rec)
        else: rec['label']=label; summ.append(rec)
    return {'accounts':acc,'summary':summ}

def parse_sojern(rows):
    # find all header rows; each starts a section named by nearest preceding non-empty label
    out=[]
    hrs=[i for i,r in enumerate(rows) if any('primary account' in txt(x).lower() for x in r) and any('yoy' in txt(x).lower() and '2027' in txt(x) for x in r)]
    for k,hi in enumerate(hrs):
        H=[txt(x) for x in rows[hi]]
        c_rank=find_col(H,lambda h:h in ('#','s no.'))
        c_acc=find_col(H,lambda h:'primary account' in h)
        c_cur=find_col(H,lambda h:'ytd fy 26-27' in h)
        c_pfy=find_col(H,lambda h:'ytd fy 25-26' in h)
        c_pfy2=find_col(H,lambda h:'ytd fy 24-25' in h)
        c_y27=find_col(H,lambda h:'2027' in h)
        c_y26=find_col(H,lambda h:'2026' in h)
        # section name = nearest non-empty label above header (not another header)
        name=''
        for i in range(hi-1,max(hi-4,-1),-1):
            t=txt(rows[i][c_acc-1]) if c_acc and c_acc-1<len(rows[i]) else ''
            t2=' '.join(txt(x) for x in rows[i] if txt(x))
            if 'Top' in t2 and 'Primary' not in t2: name=t2; break
        end=hrs[k+1] if k+1<len(hrs) else len(rows)
        acc=[]
        for row in rows[hi+1:end]:
            rank=row[c_rank] if c_rank is not None and c_rank<len(row) else None
            label=txt(row[c_acc]) if c_acc is not None and c_acc<len(row) else ''
            if not label or not isinstance(rank,(int,float)): continue
            acc.append({'rank':int(rank),'account':label,
                'ytd_cur':num(row[c_cur]) if c_cur is not None and c_cur<len(row) else None,
                'ytd_pfy':num(row[c_pfy]) if c_pfy is not None and c_pfy<len(row) else None,
                'ytd_pfy2':num(row[c_pfy2]) if c_pfy2 is not None and c_pfy2<len(row) else None,
                'yoy_2027':num(row[c_y27]) if c_y27 is not None and c_y27<len(row) else None,
                'yoy_2026':num(row[c_y26]) if c_y26 is not None and c_y26<len(row) else None})
        out.append({'section':name or ('Top' ),'accounts':acc})
    return out

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    res={'meta':{k:v for k,v in m.items() if k!='tabs'}, 'tabs':[]}
    for tab in wb.sheetnames:
        if tab=='Meta': continue
        info=m['tabs'].get(tab,{'schema':'rg','bu':tab,'currency':''})
        rows=grid(wb[tab])
        if info['schema']=='sojern':
            secs=parse_sojern(rows)
            res['tabs'].append({'tab':tab,'schema':'sojern','bu':info['bu'],'currency':info['currency'],'sections':secs})
        else:
            hi=next((i for i,r in enumerate(rows) if any('primary account' in txt(x).lower() for x in r)),None)
            parsed=parse_rg(rows,hi)
            res['tabs'].append({'tab':tab,'schema':'rg','bu':info['bu'],'currency':info['currency'],**parsed})
    return res

def validate(res):
    ok=True; msg=[]
    for tb in res['tabs']:
        if tb['schema']!='rg': continue
        summ={s['label']:s for s in tb.get('summary',[])}
        topn=next((v for k,v in summ.items() if k.lower().startswith('top')),None)
        oth=next((v for k,v in summ.items() if 'other' in k.lower()),None)
        tot=next((v for k,v in summ.items() if 'total' in k.lower()),None)
        if topn and oth and tot and None not in (topn['ytd_cur'],oth['ytd_cur'],tot['ytd_cur']):
            if abs((topn['ytd_cur']+oth['ytd_cur'])-tot['ytd_cur'])>max(1.0,0.01*abs(tot['ytd_cur'])):
                ok=False; msg.append(f"{tb['bu']}: Top+Others != Total")
    return ok,(msg or ['Top N + Others = Total holds for all RG tabs'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/Top_Accounts_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'])
    for tb in R['tabs']:
        if tb['schema']=='rg': print(f"  [rg] {tb['bu']}: {len(tb['accounts'])} accounts + {len(tb['summary'])} summary ({tb['currency']})")
        else: print(f"  [sojern] {tb['bu']}: sections="+str([(s['section'],len(s['accounts'])) for s in tb['sections']])+f" ({tb['currency']})")
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:5]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/top_accounts_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
