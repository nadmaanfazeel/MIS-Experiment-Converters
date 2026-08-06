# -*- coding: utf-8 -*-
"""SG&A Expense converter (2 tabs, YTD hierarchical, full USD).
- 'hierarchy_ytd' tab: function -> line-item, parent value = sum(children).
  Parent/child assigned by indentation + SUM-RECONCILIATION (handles un-indented children).
- 'legal_prof' tab: group -> item with a 'Nature' description column."""
import openpyxl, datetime, re, json, sys

def num(c): return round(float(c),2) if isinstance(c,(int,float)) else None
def raw(c): return c
def txt(c): return '' if c is None else str(c)
def strip(c): return txt(c).strip()
def indented(c): return len(txt(c))-len(txt(c).lstrip()) > 0
def grid(ws): return [[ws.cell(r,cc).value for cc in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

def read_meta(wb):
    m={'tabs':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=strip(row[0]) if row else ''; b=strip(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='tab' and b=='type': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['tabs'][a]=b
    return m

def conv_hierarchy(ws):
    rows=grid(ws)
    hi=next((i for i,r in enumerate(rows) if 'Particulars' in [strip(x) for x in r]),None)
    H=[strip(x) for x in rows[hi]]; pc=H.index('Particulars')
    vcols=[j for j in range(pc+1,len(H)) if H[j]]
    colnames=[H[j] for j in vcols]
    ACT_IDX=vcols[1]  # YTD FY 26-27 Act = 2nd value col
    out=[]; parent=None; psum=0.0; pval=None
    def complete(): return pval is not None and abs(psum-pval)<=max(1.0,0.01*abs(pval))
    for row in rows[hi+1:]:
        cell=row[pc] if pc<len(row) else None; label=strip(cell)
        if not label: continue
        cur=num(row[ACT_IDX]) if ACT_IDX<len(row) else None
        is_child = indented(cell) or (parent is not None and not complete())
        rec={'function': parent if is_child else label, 'line_item':label, 'is_parent': not is_child}
        for j in vcols: rec[H[j]]=num(row[j]) if j<len(row) else None
        if is_child:
            psum += (cur or 0.0)
        else:
            parent=label; pval=cur or 0.0; psum=0.0
        out.append(rec)
    return {'cols':colnames,'rows':out}

def conv_legal(ws):
    rows=grid(ws)
    hi=next((i for i,r in enumerate(rows) if 'Professional Expenses' in [strip(x) for x in r]),None)
    H=[strip(x) for x in rows[hi]]; pc=H.index('Professional Expenses')
    ncol=H.index('Nature') if 'Nature' in H else None
    vcols=[j for j in range(pc+1,len(H)) if H[j] and j!=ncol]
    out=[]; group=None
    for row in rows[hi+1:]:
        label=strip(row[pc]) if pc<len(row) else ''
        if not label: continue
        is_child=indented(row[pc])
        rec={'group': group if is_child else label,'item':label,'is_parent':not is_child,
             'nature': strip(row[ncol]) if ncol is not None and ncol<len(row) else ''}
        for j in vcols: rec[H[j]]=num(row[j]) if j<len(row) else None
        if not is_child: group=label
        out.append(rec)
    return {'cols':[H[j] for j in vcols],'rows':out}

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    res={'meta':{k:v for k,v in m.items() if k!='tabs'}}
    for tab in wb.sheetnames:
        if tab=='Meta': continue
        typ=m['tabs'].get(tab, 'legal_prof' if 'Legal' in tab else 'hierarchy_ytd')
        if typ=='legal_prof': res['legal_prof']=conv_legal(wb[tab])
        else: res['sga']=conv_hierarchy(wb[tab])
    return res

def validate(res):
    ok=True; msg=[]; sga=res.get('sga',{})
    # parent ytd_act == sum(children ytd_act)
    rows=sga.get('rows',[]); ACT='YTD FY 26-27 Act'
    i=0
    while i<len(rows):
        if rows[i]['is_parent']:
            pv=rows[i].get(ACT) or 0; s=0; j=i+1
            while j<len(rows) and not rows[j]['is_parent']: s+=(rows[j].get(ACT) or 0); j+=1
            if abs(pv-s)>max(1.0,0.01*abs(pv)) and s>0:
                ok=False; msg.append(f"{rows[i]['line_item']}: parent {round(pv,1)} != sum(children) {round(s,1)}")
            i=j
        else: i+=1
    return ok,(msg or ['parent=sum(children) holds for all groups'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/SGA_Expense_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'])
    parents=[r['line_item'] for r in R['sga']['rows'] if r['is_parent']]
    print("SG&A rows:",len(R['sga']['rows']),"| parents:",parents)
    print("Legal&Prof rows:",len(R['legal_prof']['rows']),"| groups:",[r['item'] for r in R['legal_prof']['rows'] if r['is_parent']])
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:5]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/sga_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
