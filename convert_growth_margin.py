# -*- coding: utf-8 -*-
"""Growth & Margins Snapshot converter  (Executive Summary group).

The source is a SLIDE-STYLE snapshot, not a numeric grid: one metric CARD per
column (row 2 = card header), 1-3 free-text lines beneath each (e.g. "MarTech: 84%",
"Gross Revenue: $83,194k", "163% Y-o-Y", "$4.62mn (till 20th Jul-26)").

Design (mirrors the project pattern):
  * Parse-by-LABEL, never by fixed cell position. Cards resolve via Meta card map;
    columns roll / reorder freely as long as the header wording is kept.
  * Meta drives everything: month tokens, currency, card->key/unit, the source->dashboard
    mapping (dash_card / dash_slot / dash_order / colour) and derived cards.
  * MIXED SCALE is captured PER VALUE ($'000 'k' vs $mn 'mn') -> the #1 scale-bug source.
  * Percentages stored as FRACTIONS (0.84), formatted as % on display.
  * value_norm = normalised absolute USD for currency, fraction for percent, int for count.
  * validate() checks folder identities (Revenue-Mix sums to 100%, money-never-%,
    month-token match, every card parsed) BEFORE publish.

Outputs:  growth_margin_STANDARDIZED.json   (machine)
          Growth_Margin_STANDARDIZED_Jun26.xlsx  (human review: Meta / Standardized / Dashboard_Preview)
"""
import openpyxl, json, sys, re

# ---------- helpers ----------
def txt(c):  return '' if c is None else str(c).replace('\xa0', ' ').strip()
def norm(s): return re.sub(r'\s+', ' ', txt(s)).lower()
def grid(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]

SCALE = {'k': 1_000, 'mn': 1_000_000, 'bn': 1_000_000_000}

def parse_line(text):
    """Turn one free-text line into structured tokens.
       Returns dict(raw,label,value,unit,scale,value_norm,qualifier,note)."""
    raw = txt(text)
    if not raw:
        return None
    note = None
    # trailing / inline parenthetical -> note   e.g. "(Jun 2026)" "(incl. Sojern)"
    m = re.search(r'\(([^)]*)\)', raw)
    if m:
        note = m.group(1).strip()
    body = re.sub(r'\s*\([^)]*\)', '', raw).strip()          # strip parentheticals
    star = body.endswith('*') or raw.strip().startswith('*')
    body = body.rstrip('*').strip()
    # label : rest   (label only counts if it precedes the number)
    label = None
    if ':' in body:
        left, right = body.split(':', 1)
        # a colon that sits before the numeric payload -> it's a label
        if re.search(r'[\d$]', right):
            label, body = left.strip(), right.strip()
    value = unit = scale = value_norm = None
    mc = re.search(r'\$\s*([\d,]+(?:\.\d+)?)\s*(k|mn|bn)?', body, re.I)     # currency
    mp = re.search(r'(-?\d+(?:\.\d+)?)\s*%', body)                          # percent
    if mc:
        value = float(mc.group(1).replace(',', ''))
        scale = (mc.group(2) or '').lower() or None
        unit  = 'currency'
        value_norm = value * SCALE.get(scale, 1)
    elif mp:
        value = float(mp.group(1)) / 100.0                                 # store fraction
        unit  = 'percent'
        value_norm = value
    else:
        mn = re.search(r'(-?[\d,]+(?:\.\d+)?)', body)                      # bare count/number
        if mn:
            value = float(mn.group(1).replace(',', ''))
            unit  = 'count'
            value_norm = value
    # qualifier = descriptive words left after removing the number token
    qual = body
    for pat in (r'\$\s*[\d,]+(?:\.\d+)?\s*(?:k|mn|bn)?', r'-?\d+(?:\.\d+)?\s*%', r'-?[\d,]+(?:\.\d+)?'):
        qual = re.sub(pat, '', qual, count=1, flags=re.I)
    qual = re.sub(r'\s+', ' ', qual).strip(' :-·|').strip() or None
    return {'raw': raw, 'label': label, 'value': value, 'unit': unit,
            'scale': scale, 'value_norm': value_norm, 'qualifier': qual,
            'note': note, 'footnote_mark': star}

# ---------- Meta ----------
def read_meta(wb):
    m = {'cards': [], 'derived': []}
    if 'Meta' not in wb.sheetnames:
        return m
    mode = 'kv'
    for row in grid(wb['Meta']):
        cells = [txt(x) for x in row] + ['', '', '', '', '', '', '']
        a = cells[0]
        if a == '' and all(c == '' for c in cells[:3]):
            continue
        if a == 'key' and cells[1] == 'value':      mode = 'kv';      continue
        if a == 'card' and cells[1] == 'key':        mode = 'cards';   continue
        if a == 'derived' and cells[1] == 'dash_card': mode = 'derived'; continue
        if mode == 'kv':
            if a in ('Meta — Growth & Margins Snapshot',) or a.lower().startswith('do not'):
                continue
            if cells[1] != '':
                m[a] = cells[1]
        elif mode == 'cards' and a:
            m['cards'].append({'card': a, 'key': cells[1], 'unit': cells[2],
                               'dash_card': cells[3], 'dash_slot': cells[4],
                               'dash_order': cells[5], 'color': cells[6]})
        elif mode == 'derived' and a:
            m['derived'].append({'derived': a, 'dash_card': cells[1], 'source_key': cells[2],
                                 'from_line': cells[3], 'dash_order': cells[4], 'color': cells[5]})
    return m

# ---------- convert ----------
def convert(path):
    wb  = openpyxl.load_workbook(path, data_only=True)
    m   = read_meta(wb)
    ws  = next(wb[t] for t in wb.sheetnames if t != 'Meta')
    rows = grid(ws)
    card_names = {norm(c['card']): c for c in m['cards']}

    # header row = the row with the most card-name matches
    def matches(r): return sum(1 for x in r if norm(x) in card_names)
    hi = max(range(len(rows)), key=lambda i: matches(rows[i]))
    header = rows[hi]
    col_of = {}                                        # key -> column index
    for j, x in enumerate(header):
        c = card_names.get(norm(x))
        if c:
            col_of[c['key']] = j
    title = next((txt(r[0]) for r in rows[:hi] if txt(r[0])), m.get('title', ''))

    cards, footnotes = [], []
    for cdef in m['cards']:
        j = col_of.get(cdef['key'])
        if j is None:
            continue
        lines = []
        for r in rows[hi + 1:]:
            v = r[j] if j < len(r) else None
            if txt(v) == '':
                continue
            pl = parse_line(v)
            if not pl:
                continue
            if pl['footnote_mark'] and pl['value'] is None:
                footnotes.append(pl['raw']); continue
            lines.append(pl)
        cards.append({**{k: cdef[k] for k in ('card', 'key', 'unit', 'dash_card',
                                              'dash_slot', 'dash_order', 'color')},
                      'lines': lines})
    # sheet-level footnotes (lines starting with '*' anywhere)
    for r in rows[hi + 1:]:
        for x in r:
            s = txt(x)
            if s.startswith('*') and s not in footnotes:
                footnotes.append(s)

    return {'meta': {k: v for k, v in m.items() if k not in ('cards', 'derived')},
            'title': title, 'cards': cards, 'derived': m['derived'],
            'footnotes': footnotes}

# ---------- dashboard preview (Meta-driven assembly of GM_SNAP cards) ----------
def fmt(pl):
    if pl is None or pl['value'] is None:
        return pl['raw'] if pl else ''
    if pl['unit'] == 'currency':
        n = pl['value']; s = pl['scale'] or ''
        num = f"{n:,.2f}".rstrip('0').rstrip('.') if s == 'mn' else f"{int(round(n)):,}"
        return f"${num}{s}"
    if pl['unit'] == 'percent':
        p = pl['value'] * 100
        return f"{p:.0f}%" if abs(p - round(p)) < 1e-9 else f"{p:.2f}".rstrip('0').rstrip('.') + '%'
    return f"{int(round(pl['value'])):,}"

MONTH_RE = re.compile(r'[A-Z][a-z]{2}-\d{2}')     # e.g. Jun-26

def dashboard_preview(res):
    by_key = {c['key']: c for c in res['cards']}
    order = {}
    for c in res['cards']:
        order.setdefault(c['dash_card'], []).append(c)
    out, seen = [], set()
    for c in res['cards']:
        dc = c['dash_card']
        if c['dash_slot'] != 'headline' or dc in seen:
            continue
        seen.add(dc)
        head, lines = c, c['lines']
        pct_lines = [l for l in lines if l['unit'] == 'percent']
        # ---- value ----
        if lines and lines[0]['unit'] != 'percent':
            v = fmt(lines[0])                                  # currency / count headline
        elif len(pct_lines) >= 2:
            v = ' / '.join(fmt(l) for l in pct_lines)          # 84% / 10% / 6%
        else:
            v = fmt(lines[0]) if lines else ''
        # ---- subtitle ----
        parts = []
        if lines:
            # month token embedded in the headline label (e.g. 'Jun-26 EBITDA')
            mtok = MONTH_RE.search(lines[0]['label'] or '')
            if mtok:
                parts.append(mtok.group(0))
            # first-line note (e.g. 'till 20th Jul-26', 'incl. Sojern')
            if lines[0]['note']:
                parts.append(lines[0]['note'])
        if v.count('/') and pct_lines:                          # percent-list card -> label the slices
            parts = [' · '.join(l['label'] for l in pct_lines if l['label'])]
        else:
            for l in lines[1:]:
                if l['value'] is not None:
                    seg = ' '.join(x for x in [l['label'], fmt(l), l['qualifier']] if x)
                    parts.append(seg)
                elif l['note']:
                    parts.append(l['note'])
        s = ' · '.join(p for p in parts if p)
        # ---- extra slot (e.g. ARR -> Revenue Growth) ----
        x = ''
        for c2 in order[dc]:
            if c2['dash_slot'] == 'extra' and c2['lines']:
                x = f"{c2['card']} {fmt(c2['lines'][0])}"
        out.append({'order': int(head['dash_order']) if str(head['dash_order']).isdigit() else 99,
                    't': dc, 'v': v, 's': s, 'x': x, 'color': head['color']})
    # derived cards
    for d in res['derived']:
        src = by_key.get(d['source_key'])
        if not src:
            continue
        pct = next((pl for pl in src['lines'] if pl['unit'] == 'percent'), None)
        out.append({'order': int(d['dash_order']) if str(d['dash_order']).isdigit() else 99,
                    't': d['dash_card'], 'v': fmt(pct) if pct else '',
                    's': (pct['qualifier'] or '') if pct else '', 'x': '', 'color': d['color']})
    out.sort(key=lambda r: r['order'])
    return out

# ---------- validate ----------
def validate(res):
    ok, msg = True, []
    n = len(res['cards'])
    if n < 8:
        ok = False; msg.append(f"only {n}/8 source cards found")
    # every card parsed something
    for c in res['cards']:
        if not c['lines'] or all(l['value'] is None for l in c['lines']):
            if c['unit'] != 'text':
                ok = False; msg.append(f"{c['card']}: no parsed value")
    # identity: Revenue Mix percentages sum to ~100%
    rm = next((c for c in res['cards'] if c['key'] == 'revenue_mix'), None)
    if rm:
        tot = sum(l['value'] for l in rm['lines'] if l['unit'] == 'percent' and l['value'] is not None)
        if abs(tot - 1.0) > 0.01:
            ok = False; msg.append(f"Revenue Mix sums to {tot*100:.0f}% (expected 100%)")
        else:
            msg.append(f"Revenue Mix ties to {tot*100:.0f}%")
    # money-never-%: currency-unit cards must not have parsed a percent as their headline value
    for c in res['cards']:
        if c['unit'] == 'currency' and c['lines'] and c['lines'][0]['unit'] == 'percent':
            ok = False; msg.append(f"{c['card']}: currency card parsed a % — scale/format bug")
    # month-token consistency
    mon = norm(res['meta'].get('month', ''))
    eb = next((c for c in res['cards'] if c['key'] == 'ebitda'), None)
    if mon and eb and not any(mon in norm(l['raw']) for l in eb['lines']):
        msg.append(f"note: EBITDA line does not mention Meta month '{res['meta'].get('month')}'")
    return ok, (msg or ['all cards parsed; identities hold'])

# ---------- review Excel ----------
def write_xlsx(res, prev, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); wb.remove(wb.active)
    AR = Font(name='Arial', size=10); BD = Font(name='Arial', size=10, bold=True)
    HF = PatternFill('solid', fgColor='4B2E83'); HW = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9D9D9'); BB = Border(thin, thin, thin, thin)
    def hdr(ws, r, vals):
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.font = HW; c.fill = HF; c.border = BB
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    def cell(ws, r, cc, v, font=AR, nfmt=None, wrap=False):
        c = ws.cell(r, cc, v); c.font = font; c.border = BB
        if nfmt: c.number_format = nfmt
        if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
        return c

    # Sheet 1: Meta echo
    ws = wb.create_sheet('Meta')
    ws.cell(1, 1, 'Meta').font = BD
    r = 2
    for k, v in res['meta'].items():
        cell(ws, r, 1, k, BD); cell(ws, r, 2, v); r += 1
    r += 1; hdr(ws, r, ['card', 'key', 'unit', 'dash_card', 'dash_slot', 'dash_order', 'color']); r += 1
    for c in res['cards']:
        for j, key in enumerate(['card','key','unit','dash_card','dash_slot','dash_order','color'], 1):
            cell(ws, r, j, c[key]); r_last = r
        r += 1
    for w, col in zip([20,18,10,20,11,11,10], 'ABCDEFG'):
        ws.column_dimensions[col].width = w

    # Sheet 2: Standardized long form
    ws = wb.create_sheet('Standardized')
    cols = ['Card','Key','Card Unit','Dash Card','Slot','Line #','Label','Value',
            'Value Unit','Scale','Value (norm USD/frac)','Qualifier','Note','Raw Text']
    hdr(ws, 1, cols); r = 2
    for c in res['cards']:
        for i, l in enumerate(c['lines'], 1):
            cell(ws, r, 1, c['card']); cell(ws, r, 2, c['key']); cell(ws, r, 3, c['unit'])
            cell(ws, r, 4, c['dash_card']); cell(ws, r, 5, c['dash_slot']); cell(ws, r, 6, i)
            cell(ws, r, 7, l['label'])
            nf = '0.##%' if l['unit'] == 'percent' else ('$#,##0.00' if (l['unit']=='currency' and l['scale']=='mn')
                    else ('$#,##0' if l['unit']=='currency' else '#,##0'))
            cell(ws, r, 8, l['value'], nfmt=nf)
            cell(ws, r, 9, l['unit']); cell(ws, r, 10, l['scale'] or '')
            nfn = '0.##%' if l['unit']=='percent' else '#,##0'
            cell(ws, r, 11, l['value_norm'], nfmt=nfn)
            cell(ws, r, 12, l['qualifier']); cell(ws, r, 13, l['note'])
            cell(ws, r, 14, l['raw'], wrap=True)
            r += 1
    if res['footnotes']:
        r += 1; cell(ws, r, 1, 'Footnotes', BD)
        for f in res['footnotes']:
            r += 1; cell(ws, r, 1, f, wrap=True)
    for w, col in zip([16,15,9,16,9,7,14,12,10,7,16,20,16,34], 'ABCDEFGHIJKLMN'):
        ws.column_dimensions[col].width = w

    # Sheet 3: Dashboard preview (GM_SNAP cards, Meta-assembled)
    ws = wb.create_sheet('Dashboard_Preview')
    ws.cell(1, 1, 'Auto-assembled from the standardized data (final string polish = the map→load step). Should mirror the dashboard GM_SNAP cards.').font = AR
    hdr(ws, 3, ['#', 'Card (t)', 'Value (v)', 'Subtitle (s)', 'Extra (x)', 'Colour']); r = 4
    for p in prev:
        cell(ws, r, 1, p['order']); cell(ws, r, 2, p['t'], BD); cell(ws, r, 3, p['v'])
        cell(ws, r, 4, p['s'], wrap=True); cell(ws, r, 5, p['x']); cd = cell(ws, r, 6, p['color'])
        try: cd.fill = PatternFill('solid', fgColor=p['color'].lstrip('#'))
        except Exception: pass
        r += 1
    for w, col in zip([4,20,16,34,20,12], 'ABCDEF'):
        ws.column_dimensions[col].width = w

    wb.save(path)

# ---------- main ----------
if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs/Growth_Margin_Jun26_INPUT_withMeta.xlsx'
    R = convert(path)
    P = dashboard_preview(R)
    ok, msg = validate(R)
    print("META:", R['meta'])
    print("TITLE:", R['title'])
    for c in R['cards']:
        print(f"  [{c['unit']:8}] {c['card']}  -> dash:{c['dash_card']}/{c['dash_slot']}")
        for l in c['lines']:
            print(f"        {l['raw']!r:45} | label={l['label']!r} value={l['value']} unit={l['unit']} scale={l['scale']} norm={l['value_norm']} qual={l['qualifier']!r} note={l['note']!r}")
    print("\nDASHBOARD PREVIEW (GM_SNAP):")
    for p in P:
        print(f"  {p['order']}. {p['t']:18} v={p['v']:12} s={p['s']!r}  x={p['x']!r}")
    print("\nFOOTNOTES:", R['footnotes'])
    print("VALIDATE:", ("PASS" if ok else "FAIL"), "|", '; '.join(msg))
    json.dump({'standardized': R, 'dashboard_preview': P, 'validate': {'pass': ok, 'messages': msg}},
              open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/growth_margin_STANDARDIZED.json', 'w'),
              ensure_ascii=False, indent=1)
    write_xlsx(R, P, (__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/Growth_Margin_STANDARDIZED_Jun26.xlsx')
    print("\nWrote growth_margin_STANDARDIZED.json + Growth_Margin_STANDARDIZED_Jun26.xlsx")
