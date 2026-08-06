# -*- coding: utf-8 -*-
"""ONE converter for the whole Full P&L workbook (3 data tabs + Meta):
  1. DaaS+Distribution+SoHo         -> product line-item P&L
  2. Sojern - Consolidated          -> consolidated block (incl. all growth/var %)
  3. Property+Corporate+Destination -> 3 entity blocks (incl. Δ%)
Reads the Meta tab for month tokens (auto-detect fallback). Roles are FIXED so the
rolling month window resolves automatically each month. Keeps every % column."""
import openpyxl, datetime, re, json, sys

def num(c): return round(float(c),4) if isinstance(c,(int,float)) else None
def txt(c): return '' if c is None else (c.strftime('%b-%y') if isinstance(c,datetime.datetime) else str(c)).strip()

LBL={'GR':'Gross Revenue','NR':'Net Revenue','Gross Revenue (GR)':'Gross Revenue',
     'Net Revenue (NR)':'Net Revenue','EBITDA $':'EBITDA','Total Operating cost':'Total Operating Cost'}
def canon(l): l=re.sub(r'\s+',' ',l).strip(); return LBL.get(l,l)

def read_meta(wb):
    d={}
    if 'Meta' in wb.sheetnames:
        for r in range(1,wb['Meta'].max_row+1):
            k=txt(wb['Meta'].cell(r,1).value); v=txt(wb['Meta'].cell(r,2).value)
            if k and k.lower()!='key' and not k.endswith(':'): d[k]=v
    if d.get('month'):
        mm=datetime.datetime.strptime(d['month'],'%b-%y')
        d['cur_mon']=mm.strftime('%b'); d['cur_yy']=mm.strftime('%y')
        d['py_yy']=(d.get('prior_year_month','') or '')[-2:] or '%02d'%(mm.year-1)%100
    return d

def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

# ---------- Tab 1: DaaS+Distribution+SoHo ----------
MONTH_ACT=re.compile(r"^([A-Za-z]{3})-(\d{2}) Act$")
def parse_mon(s):
    m=MONTH_ACT.match(txt(s)); return datetime.datetime.strptime(f"{m.group(1)}-{m.group(2)}","%b-%y") if m else None
def clean_prod_label(product,raw):
    lbl=re.sub(r'\s+',' ',txt(raw)); w=lbl.split(' ')
    if len(w)>=2 and w[-1]==w[-2]: lbl=' '.join(w[:-1])
    for p in [product,product.replace('PG - ','').replace('PG-','')]:
        if lbl.startswith(p+' '): lbl=lbl[len(p)+1:]
    return lbl.strip()
def convert_products(ws,meta):
    out=[]; product=None; cm=None
    for row in grid(ws):
        label=txt(row[2]) if len(row)>2 else ''
        is_head=any(parse_mon(row[j]) for j in range(3,len(row))) and label
        if is_head:
            product=label
            cur=f"{meta['cur_mon']}-{meta['cur_yy']}"; prior=meta.get('prior_month','')
            exact={cur+' Act':'cur_act',cur+' Bud':'cur_bud',prior+' Act':'prior_act',
                   'MoM Bud Var':'mom_bud_var','YTD '+meta.get('cur_fy','')+' Act':'ytd_act',
                   'YTD '+meta.get('cur_fy','')+' Bud':'ytd_bud','YoY Bud Var':'yoy_bud_var',
                   'YTD '+meta.get('prior_fy','')+' Act':'pfy_ytd_act','YoY Var':'yoy_var',
                   f"{meta['cur_mon']}-{meta['py_yy']} Act":'py_month_act','YoY month Var %':'yoy_month_var'}
            cm={}; 
            for j in range(3,len(row)):
                h=txt(row[j])
                if h in exact: cm[exact[h]]=j
            continue
        if product and cm and label:
            rec={'product':product,'line_item':clean_prod_label(product,label)}
            for role,j in cm.items(): rec[role]=num(row[j]) if j<len(row) else None
            if rec.get('cur_act') is not None or rec.get('ytd_act') is not None: out.append(rec)
    return out

# ---------- Tab 2: Sojern - Consolidated ----------
def convert_sojern_consol(ws,meta):
    cur=meta['cur_mon']; cy=meta['cur_yy']; py=meta['py_yy']; out=[]; cm=None
    ROLES={f'Budget {cur}{cy}':'cur_bud',f'Actual {cur}{py}':'py_month_act',f'Actual {cur}{cy}':'cur_act',
           'YoY Growth $':'yoy_growth_d','YoY Growth %':'yoy_growth_pct','Growth from Budget $':'growth_from_bud_d','Growth %':'growth_from_bud_pct',
           f'Budget YTD{cy}':'ytd_bud',f'Actual YTD{py}':'pfy_ytd_act',f'Actual YTD{cy}':'ytd_act',
           'YTD YoY Growth $':'ytd_yoy_growth_d','YTD YoY Growth %':'ytd_yoy_growth_pct','YTD Growth from Bud $':'ytd_growth_from_bud_d','YTD Growth %':'ytd_growth_from_bud_pct'}
    for row in grid(ws):
        label=txt(row[2]) if len(row)>2 else ''
        if label.startswith('Desc.'):
            cm={ROLES[txt(row[j])]:j for j in range(3,len(row)) if txt(row[j]) in ROLES}; continue
        if cm and label and any(isinstance(x,(int,float)) for x in row[3:]):
            rec={'line_item':canon(label)}
            for role,j in cm.items(): rec[role]=num(row[j]) if j<len(row) else None
            out.append(rec)
    return out

# ---------- Tab 3: Property+Corporate+Destination ----------
def convert_sojern_entities(ws,meta):
    cur=meta['cur_mon']; out=[]; entity=None; cm=None
    ROLES={f'{cur} Bud':'cur_bud','YTD Bud':'ytd_bud',f'{cur} Act':'cur_act','YTD Act':'ytd_act',f'{cur} \u0394%':'mom_delta_pct','YTD \u0394%':'ytd_delta_pct'}
    for row in grid(ws):
        label=txt(row[2]) if len(row)>2 else ''
        has_nums=any(isinstance(x,(int,float)) for x in row[3:])
        if label.startswith('Sojern -') and not has_nums:
            entity=label.replace('Sojern -','').strip(); cm=None; continue
        if label=='Particulars':
            cm={ROLES[txt(row[j])]:j for j in range(3,len(row)) if txt(row[j]) in ROLES}; continue
        if entity and cm and label and has_nums:
            rec={'entity':entity,'line_item':canon(label)}
            for role,j in cm.items(): rec[role]=num(row[j]) if j<len(row) else None
            out.append(rec)
    return out

def convert_all(path):
    wb=openpyxl.load_workbook(path,data_only=True); meta=read_meta(wb)
    res={'meta':{k:meta[k] for k in ('month','cur_fy','prior_fy','prior_month','prior_year_month','currency') if k in meta}}
    if 'DaaS+Distribution+SoHo' in wb.sheetnames: res['daas_dist_soho']=convert_products(wb['DaaS+Distribution+SoHo'],meta)
    if 'Sojern - Consolidated' in wb.sheetnames: res['sojern_consolidated']=convert_sojern_consol(wb['Sojern - Consolidated'],meta)
    if 'Property+Corporate+Destination' in wb.sheetnames: res['sojern_entities']=convert_sojern_entities(wb['Property+Corporate+Destination'],meta)
    return res

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/uploads/Full_P_L_Jun_26.xlsx'
    R=convert_all(path)
    print("META:",R['meta'])
    print("daas_dist_soho rows:",len(R.get('daas_dist_soho',[])),"| products:",len({r['product'] for r in R.get('daas_dist_soho',[])}))
    print("sojern_consolidated rows:",len(R.get('sojern_consolidated',[])))
    print("sojern_entities rows:",len(R.get('sojern_entities',[])),"| entities:",sorted({r['entity'] for r in R.get('sojern_entities',[])}))
    print("\n--- Sojern-Consolidated 'GR' (ALL columns incl %) ---")
    print(json.dumps([r for r in R['sojern_consolidated'] if r['line_item']=='Gross Revenue'][0],ensure_ascii=False,indent=1))
    print("\n--- Property 'EBITDA' (incl Δ%) ---")
    print(json.dumps([r for r in R['sojern_entities'] if r['entity']=='Property' and r['line_item']=='EBITDA'][0],ensure_ascii=False))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/full_pnl_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
