# -*- coding: utf-8 -*-
"""HR headcount converter (7 tabs). Meta-driven tab types:
- 'flat'      : Department -> count (+ Total)
- 'hierarchy' : Region/Entity -> Category(FTE/Intern/Part-time) -> count (+ Total)
Point-in-time snapshot; the value column header is the month."""
import openpyxl, datetime, json, sys

def num(c): return int(round(c)) if isinstance(c,(int,float)) else None
def txt(c): return '' if c is None else (c.strftime("%b-%y") if isinstance(c,datetime.datetime) else str(c)).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

def read_meta(wb):
    m={'tabs':{},'categories':[]}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='tab' and b=='type': mode='map'; continue
        if mode=='kv' and a.lower()!='key':
            if a=='categories': m['categories']=[x.strip() for x in b.split('|')]
            else: m[a]=b
        elif mode=='map': m['tabs'][a]=b
    return m

def conv_flat(ws):
    rows=grid(ws); month=txt(rows[0][1]) if len(rows[0])>1 else ''
    out=[]
    for row in rows[1:]:
        lab=txt(row[0]); v=num(row[1]) if len(row)>1 else None
        if not lab: continue
        out.append({'label':lab,'count':v,'is_total':lab.lower()=='total'})
    return {'type':'flat','month':month,'rows':out}

def conv_hier(ws, cats):
    rows=grid(ws); month=txt(rows[0][1]) if len(rows[0])>1 else ''
    out=[]; group=None
    catset=set(cats)
    for row in rows[1:]:
        lab=txt(row[0]); v=num(row[1]) if len(row)>1 else None
        if not lab: continue
        if lab.lower()=='total': out.append({'node':'Total','level':0,'parent':None,'count':v,'is_total':True}); continue
        if lab in catset: out.append({'node':lab,'level':1,'parent':group,'count':v,'is_total':False})
        else: group=lab; out.append({'node':lab,'level':0,'parent':None,'count':v,'is_total':False})
    return {'type':'hierarchy','month':month,'rows':out}

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    cats=m.get('categories') or ['FTE','Intern/Apprentice','Part Time/Contractor/Rainmaker']
    res={'meta':{k:v for k,v in m.items() if k not in ('tabs',)}, 'tables':{}}
    for tab in wb.sheetnames:
        if tab=='Meta': continue
        typ=m['tabs'].get(tab.strip(), m['tabs'].get(tab,'flat'))
        res['tables'][tab.strip()] = conv_hier(wb[tab],cats) if typ=='hierarchy' else conv_flat(wb[tab])
    return res

def validate(res):
    ok=True; msg=[]; T=res['tables']
    # within-tab totals
    for name,tb in T.items():
        rows=tb['rows']
        if tb['type']=='flat':
            tot=next((r['count'] for r in rows if r['is_total']),None)
            s=sum(r['count'] or 0 for r in rows if not r['is_total'])
            if tot is not None and s!=tot: ok=False; msg.append(f"{name}: sum {s} != Total {tot}")
        else:
            # each group == sum of its categories
            groups=[r for r in rows if r['level']==0 and not r['is_total']]
            for g in groups:
                s=sum(r['count'] or 0 for r in rows if r['parent']==g['node'])
                if g['count'] is not None and s!=g['count']: ok=False; msg.append(f"{name}/{g['node']}: cats {s} != {g['count']}")
            tot=next((r['count'] for r in rows if r['is_total']),None)
            gs=sum(g['count'] or 0 for g in groups)
            if tot is not None and gs!=tot: ok=False; msg.append(f"{name}: groups {gs} != Total {tot}")
    # cross-tab: Combine entities == individual tab totals; Combine Total == By Region Total
    def tab_total(nm): 
        tb=T.get(nm); 
        return next((r['count'] for r in tb['rows'] if r.get('is_total')),None) if tb else None
    comb=T.get('Combine_SoHo_Sojern_RG')
    if comb:
        cmap={r['node']:r['count'] for r in comb['rows'] if r['level']==0}
        for tabname,combname in [('RG','Rest of RG'),('SoHo','Soho'),('Sojern','Sojern')]:
            tt=tab_total(tabname); cc=cmap.get(combname)
            if tt is not None and cc is not None and tt!=cc: ok=False; msg.append(f"cross: {tabname} total {tt} != Combine '{combname}' {cc}")
        if tab_total('Combine_SoHo_Sojern_RG') is not None and tab_total('Total Headcount By Region') is not None:
            if tab_total('Combine_SoHo_Sojern_RG')!=tab_total('Total Headcount By Region'):
                ok=False; msg.append("cross: Combine Total != By Region Total")
    return ok,(msg or ['within-tab + cross-tab headcount reconciliations hold'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/HR_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'])
    for nm,tb in R['tables'].items(): print(f"  {nm}: {tb['type']} ({len(tb['rows'])} rows, month {tb['month']})")
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:6]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/hr_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
