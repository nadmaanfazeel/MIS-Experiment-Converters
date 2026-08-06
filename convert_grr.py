# -*- coding: utf-8 -*-
"""GRR/NRR Ratios converter. Single matrix: entities as COLUMNS, bridge+ratio rows.
Transposes to one tidy record per entity. Meta supplies entity->type/bu.
Revenue rows in $'000; GRR/NRR as fractions."""
import openpyxl, datetime, json, sys

def num(c): return round(float(c),3) if isinstance(c,(int,float)) else None
def txt(c): return '' if c is None else (str(c)).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

ROWMAP={'YTD FY 25-26':'ytd_pfy','Churn':'churn','Downsell':'downsell','Upsell':'upsell',
        'New Revenue':'new_revenue','Exceptional Items':'exceptional','YTD FY 26-27':'ytd_cur',
        'GRR':'grr','NRR':'nrr'}

def read_meta(wb):
    m={'entities':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='entity' and b=='type': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['entities'][a]={'type':b,'bu':txt(row[2]) if len(row)>2 else ''}
    return m

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    ws=next(wb[t] for t in wb.sheetnames if t!='Meta'); rows=grid(ws)
    hi=next(i for i,r in enumerate(rows) if any('Particulars' in txt(x) for x in r))
    H=[txt(x) for x in rows[hi]]; lc=next(j for j,h in enumerate(H) if 'Particulars' in h)
    ent_cols=[(j,H[j]) for j in range(lc+1,len(H)) if H[j]]
    recs={e:{'entity':e} for j,e in ent_cols}
    for row in rows[hi+1:]:
        lbl=txt(row[lc]) if lc<len(row) else ''
        if lbl not in ROWMAP: continue
        field=ROWMAP[lbl]
        for j,e in ent_cols: recs[e][field]=num(row[j]) if j<len(row) else None
    out=[]
    for j,e in ent_cols:
        r=recs[e]; info=m['entities'].get(e,{'type':'','bu':''})
        r['type']=info['type']; r['bu']=info['bu']; out.append(r)
    return {'meta':{k:v for k,v in m.items() if k!='entities'}, 'entities':out}

def validate(res):
    ok=True; msg=[]
    idx={r['entity']:r for r in res['entities']}
    # 1) bridge identity per entity
    for r in res['entities']:
        comp=[r.get(k) for k in ('ytd_pfy','churn','downsell','upsell','new_revenue','exceptional')]
        if all(v is not None for v in comp) and r.get('ytd_cur') is not None:
            if abs(sum(comp)-r['ytd_cur'])>max(1.0,0.01*abs(r['ytd_cur'])):
                ok=False; msg.append(f"{r['entity']}: bridge {round(sum(comp),1)} != YTD FY26-27 {r['ytd_cur']}")
    # 2) subtotal identities
    def s(*names): return sum((idx[n]['ytd_cur'] or 0) for n in names if n in idx)
    checks=[('DaaS',['TravelBI','HospiBI']),('Distribution',['UNO+RZ','Ent. Conn.']),
            ('Consol',['DaaS','Distribution','SoHo'])]
    for tot,parts in checks:
        if tot in idx and all(p in idx for p in parts):
            if abs((idx[tot]['ytd_cur'] or 0)-s(*parts))>max(1.0,0.01*abs(idx[tot]['ytd_cur'] or 1)):
                ok=False; msg.append(f"{tot} != sum({parts})")
    return ok,(msg or ['bridge + subtotal identities hold'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/GRR_NRR_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'])
    print("entities:",[r['entity'] for r in R['entities']])
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:6]))
    ex=[r for r in R['entities'] if r['entity']=='PG-OTA'][0]
    print("PG-OTA:",json.dumps(ex,ensure_ascii=False))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/grr_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
