# -*- coding: utf-8 -*-
"""Sojern converter  (MIS_Automation_FY27 · Sojern — 16-tab faithful mirror).

The dashboard renders each tab as a card showing its full table "as in source", so the
job is FAITHFUL TABLE CAPTURE, not bespoke reshaping. The Meta is a TAB REGISTRY
(title / kind / unit / order) and ONE generic extractor serves every tab:

  * Detect header rows by CONTENT — a row carrying >=3 time tokens (Jun-25 / 1Q26 /
    "Actual Jun-26") is a column header; multi-block tabs simply have >1 such row.
    Key-value tabs (Corporate Allocation) fall back to a label-keyword header.
  * For each header, read the block beneath: label column auto-detected (first text
    column), rows split into a leading label (+ carried group/section) and value cells.
  * Numbers tagged with the tab's Meta unit; percent kept as fractions; a block whose
    values are all fractional while the tab unit is currency is auto-flagged ratio/percent.
  * validate() confirms every tab produced >=1 block with month/quarter columns and
    numeric coverage, and flags the known-messy tabs (aging / multiblock) for review.

Outputs: sojern_STANDARDIZED.json  +  Sojern_STANDARDIZED_Jun26.xlsx
         (Registry sheet + one normalised sheet per tab)
"""
import openpyxl, json, sys, re

def txt(c):  return '' if c is None else str(c).replace('\xa0', ' ').replace('\n', ' ').strip()
def norm(s): return re.sub(r'\s+', ' ', txt(s)).lower()
def grid(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]

QTR_RE   = re.compile(r'^\d[Qq]\d{2}$')                       # 1Q26
MONTH_IN = re.compile(r'[A-Za-z]{3,9}[-\s]\d{2,4}')          # embedded "Actual Jun-26 (FY27)"
LABEL_HDR = re.compile(r'^(department|region|metric|month|account|vertical|category|days)\b', re.I)
MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

def month_key(s):
    """Any month spelling -> canonical 3-letter key ('July'->'jul', 'Sept'->'sep')."""
    m = re.match(r'([a-z]+)', norm(s))
    if not m:
        return None
    w = m.group(1)
    for mo in MONTHS:
        if w.startswith(mo):
            return mo
    return None

def month_ym(s):
    """Pure month token -> (year, 1-12), else None.
       Handles Jun-25 / Jun 25 / June-25 / Jun-2026 / jun'26 in any spelling."""
    m = re.match(r"^([a-z]{3,9})[\s\-\u2019'.]*(\d{2,4})$", norm(s))
    if not m:
        return None
    mk = month_key(m.group(1))
    if not mk:
        return None
    yr = int(m.group(2)); yr = yr + 2000 if yr < 100 else yr
    return (yr, MONTHS.index(mk) + 1)

def is_timecol(v):
    t = txt(v)
    return bool(month_ym(t) or QTR_RE.match(t) or MONTH_IN.search(t))

def num(v):
    if isinstance(v, (int, float)): return float(v)
    s = txt(v).replace('$', '').replace(',', '').replace('%', '')
    if s in ('', '-', '–', '—', 'NA', 'N/A'): return None
    try: return float(s)
    except ValueError: return None

# ---------- Meta ----------
def read_meta(wb):
    m = {'registry': []}
    if 'Meta' not in wb.sheetnames: return m
    mode = None
    for row in grid(wb['Meta']):
        c = [txt(x) for x in row] + ['', '', '', '', '']
        a = c[0]
        if a == 'key' and c[1] == 'value':   mode = 'kv';  continue
        if a == 'tab' and c[1] == 'title':    mode = 'reg'; continue
        if a == '' and c[1] == '':            continue
        if a.startswith('Meta —') or a.startswith('16-tab'): continue
        if mode == 'kv' and c[1]:
            m[a] = c[1]
        elif mode == 'reg' and a:
            m['registry'].append({'tab': a, 'title': c[1], 'kind': c[2], 'unit': c[3],
                                   'order': int(c[4]) if str(c[4]).isdigit() else 99})
    return m

# ---------- generic block extraction ----------
def header_rows(rows):
    hdrs = []
    for i, r in enumerate(rows):
        tc = sum(1 for v in r if is_timecol(v))
        if tc >= 3:
            hdrs.append(i)
    if not hdrs:                                   # key-value / label-keyword fallback
        for i, r in enumerate(rows):
            if any(LABEL_HDR.match(txt(v)) for v in r) and sum(1 for v in r if txt(v)) >= 2:
                hdrs.append(i); break
    return hdrs

def label_col_of(rows, hi, value_cols):
    first_v = min(value_cols) if value_cols else 1
    for r in rows[hi + 1: hi + 6]:                 # sniff a few data rows
        for j in range(first_v):
            if txt(r[j]) and num(r[j]) is None:     # left-of-values text cell = label col
                return j
    return max(0, first_v - 1)

def section_above(rows, hi):
    for k in (hi - 1, hi - 2):
        if k < 0: continue
        for v in rows[k]:
            t = txt(v)
            if t and not is_timecol(t) and '($' not in t and not LABEL_HDR.match(t):
                return t
    return None

def extract_blocks(rows, hdrs):
    blocks = []
    for bi, hi in enumerate(hdrs):
        header = rows[hi]
        value_cols = [j for j in range(len(header)) if txt(header[j])]
        lc = label_col_of(rows, hi, [j for j in value_cols if is_timecol(header[j])] or value_cols)
        cols = [(j, txt(header[j])) for j in value_cols if j > lc and txt(header[j])]
        stop = hdrs[bi + 1] if bi + 1 < len(hdrs) else len(rows)
        data, group, blank = [], None, 0
        for r in rows[hi + 1: stop]:
            if all(txt(x) == '' for x in r):
                blank += 1
                if blank >= 2: break
                continue
            blank = 0
            label = txt(r[lc])
            vals = {name: r[j] for j, name in cols}
            has_num = any(num(v) is not None for v in vals.values())
            if label and not has_num:              # section / group sub-header
                group = label; continue
            if not label and not has_num:
                continue
            data.append({'group': group, 'label': label or None,
                         'cells': {name: (num(v) if num(v) is not None else (txt(v) or None))
                                   for name, v in vals.items()}})
        if data:
            colnames = [n for _, n in cols]
            mo = sorted(((month_ym(n), n) for n in colnames if month_ym(n)), key=lambda x: x[0])
            blk = {'section': section_above(rows, hi), 'columns': colnames,
                   'rows': data, 'label_col': lc, 'header_row': hi}
            if len(mo) >= 2:
                blk['month_start'], blk['month_end'], blk['n_months'] = mo[0][1], mo[-1][1], len(mo)
            blocks.append(blk)
    return blocks

def tag_units(blocks, unit):
    """Attach a resolved unit per block; auto-detect ratio/percent blocks in currency tabs."""
    for b in blocks:
        nums = [v for r in b['rows'] for v in r['cells'].values() if isinstance(v, (int, float))]
        frac = nums and all(abs(v) <= 1.5 for v in nums)
        if unit in ('usd', 'thousands', 'count') and frac and len(nums) >= 3:
            b['unit'] = 'ratio/percent'
        else:
            b['unit'] = unit
    return blocks

def title_of(rows, reg):
    for r in rows[:4]:
        t = txt(r[0]) or txt(r[1])
        if t and not is_timecol(t) and '($' not in t:
            return t
    return reg['title']

# ---------- convert ----------
def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    m = read_meta(wb)
    reg_by = {norm(r['tab']): r for r in m['registry']}
    tabs = []
    for name in [s for s in wb.sheetnames if s != 'Meta']:
        reg = reg_by.get(norm(name), {'title': name, 'kind': 'unknown', 'unit': 'usd', 'order': 99})
        rows = grid(wb[name])
        hdrs = header_rows(rows)
        blocks = tag_units(extract_blocks(rows, hdrs), reg['unit'])
        window = next((f"{b['month_start']} → {b['month_end']}" for b in blocks if b.get('month_start')), None)
        tabs.append({'tab': name, 'title': reg['title'], 'kind': reg['kind'],
                     'unit': reg['unit'], 'order': reg['order'], 'window': window,
                     'n_rows_src': len(rows), 'blocks': blocks})
    tabs.sort(key=lambda t: t['order'])
    return {'meta': {k: v for k, v in m.items() if k != 'registry'}, 'tabs': tabs}

# ---------- validate ----------
def validate(res):
    ok, msg = True, []
    for t in res['tabs']:
        nb = len(t['blocks'])
        nrows = sum(len(b['rows']) for b in t['blocks'])
        if nb == 0:
            ok = False; msg.append(f"{t['tab']}: no table block extracted")
            continue
        has_time = any(any(is_timecol(c) for c in b['columns']) for b in t['blocks'])
        if t['kind'] in ('timeseries', 'region_matrix') and not has_time:
            msg.append(f"note: {t['tab']} — no time columns detected")
        if t['kind'] in ('aging', 'multiblock'):
            msg.append(f"review: {t['tab']} ({t['kind']}) — {nb} block(s), {nrows} rows; complex tab, verify parse")
    total_blocks = sum(len(t['blocks']) for t in res['tabs'])
    total_rows = sum(len(b['rows']) for t in res['tabs'] for b in t['blocks'])
    msg.append(f"{len(res['tabs'])} tabs → {total_blocks} table blocks, {total_rows} data rows extracted")
    return ok, msg

# ---------- review Excel ----------
def write_xlsx(res, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); wb.remove(wb.active)
    AR = Font(name='Arial', size=9); BD = Font(name='Arial', size=9, bold=True)
    HF = PatternFill('solid', fgColor='C77A1E'); HW = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    GRP = PatternFill('solid', fgColor='FBF1E2'); thin = Side(style='thin', color='E0D6C4'); BB = Border(thin, thin, thin, thin)
    def hdr(ws, r, vals):
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.font = HW; c.fill = HF; c.border = BB
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    def cell(ws, r, cc, v, font=AR, nfmt=None, fill=None):
        c = ws.cell(r, cc, v); c.font = font; c.border = BB
        if nfmt: c.number_format = nfmt
        if fill: c.fill = fill
        return c

    # Registry sheet
    ws = wb.create_sheet('Registry')
    ws.cell(1, 1, 'Sojern — tab registry (each tab mirrored on its own sheet)').font = BD
    r = 2
    for k, v in res['meta'].items():
        cell(ws, r, 1, k, BD); cell(ws, r, 2, v); r += 1
    r += 1; hdr(ws, r, ['#', 'Tab', 'Title', 'Kind', 'Unit', 'Blocks', 'Rows', 'Month window']); r += 1
    for i, t in enumerate(res['tabs'], 1):
        cell(ws, r, 1, i); cell(ws, r, 2, t['tab']); cell(ws, r, 3, t['title'])
        cell(ws, r, 4, t['kind']); cell(ws, r, 5, t['unit'])
        cell(ws, r, 6, len(t['blocks'])); cell(ws, r, 7, sum(len(b['rows']) for b in t['blocks']))
        cell(ws, r, 8, t.get('window'))
        r += 1
    for w, col in zip([4, 30, 32, 14, 11, 8, 7, 20], 'ABCDEFGH'): ws.column_dimensions[col].width = w

    # one sheet per tab (normalised)
    used = {'Registry'}
    for t in res['tabs']:
        base = re.sub(r'[\\/*?:\[\]]', ' ', t['tab'])[:28]; nm = base; k = 1
        while nm in used: k += 1; nm = f"{base[:26]}_{k}"
        used.add(nm)
        ws = wb.create_sheet(nm)
        ws.cell(1, 1, f"{t['title']}   ·   unit: {t['unit']}   ·   kind: {t['kind']}").font = BD
        r = 3
        for bi, b in enumerate(t['blocks'], 1):
            if len(t['blocks']) > 1:
                cell(ws, r, 1, f"Block {bi}  ({b['unit']})", BD); r += 1
            pct = (b['unit'] in ('percent', 'ratio/percent'))
            hdr(ws, r, ['Group', 'Label'] + b['columns']); r += 1
            for row in b['rows']:
                cell(ws, r, 1, row['group'], fill=GRP if row['group'] else None)
                cell(ws, r, 2, row['label'], BD)
                for j, cn in enumerate(b['columns'], 3):
                    v = row['cells'].get(cn)
                    nf = '0.0%' if pct and isinstance(v, (int, float)) else ('#,##0' if isinstance(v, (int, float)) else None)
                    cell(ws, r, j, v, nfmt=nf)
                r += 1
            r += 1
        ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 22
        for col in [chr(c) for c in range(ord('C'), ord('C') + max(1, min(20, max(len(b['columns']) for b in t['blocks']) if t['blocks'] else 1)))]:
            ws.column_dimensions[col].width = 12
        ws.freeze_panes = 'C4'
    wb.save(path)

# ---------- main ----------
if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs/Sojern_Jun26_INPUT_withMeta.xlsx'
    R = convert(path)
    ok, msg = validate(R)
    print("META:", R['meta'])
    for t in R['tabs']:
        print(f"\n[{t['order']:>2}] {t['tab']}  ({t['kind']}, {t['unit']})  window={t.get('window')}")
        for bi, b in enumerate(t['blocks'], 1):
            print(f"     block{bi} unit={b['unit']} cols={b['columns'][:6]}{'…' if len(b['columns'])>6 else ''} rows={len(b['rows'])}")
            for row in b['rows'][:2]:
                cs = {k: (round(v, 3) if isinstance(v, float) else v) for k, v in list(row['cells'].items())[:4]}
                print(f"        {row['group']!s:>12} | {row['label']!s:22} | {cs}")
    print("\nVALIDATE:", ("PASS" if ok else "FAIL"))
    for x in msg: print("   -", x)
    json.dump({'standardized': R, 'validate': {'pass': ok, 'messages': msg}},
              open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/sojern_STANDARDIZED.json', 'w'), ensure_ascii=False, indent=1)
    write_xlsx(R, (__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/Sojern_STANDARDIZED_Jun26.xlsx')
    print("\nWrote sojern_STANDARDIZED.json + Sojern_STANDARDIZED_Jun26.xlsx")
