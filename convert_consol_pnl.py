# -*- coding: utf-8 -*-
"""Converter for the Consolidated P&L Summary file (BU x metric grid).
Meta-driven parse-by-header -> fixed roles (keeps all variance %). Includes validate()."""
import openpyxl, datetime, re, json, sys

def num(c): return round(float(c),4) if isinstance(c,(int,float)) else None
def txt(c): return '' if c is None else (c.strftime('%b-%y') if isinstance(c,datetime.datetime) else str(c)).strip()
BUS=['DaaS','Distribution','Martech','Consolidated']

def read_meta(wb):
    d={}
    if 'Meta' in wb.sheetnames:
        for r in range(1,wb['Meta'].max_row+1):
            k=txt(wb['Meta'].cell(r,1).value); v=txt(wb['Meta'].cell(r,2).value)
            if k and k.lower()!='key' and not k.endswith(':'): d[k]=v
    return d

def convert(path, tab='Consolidated P&L Summary'):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb); ws=wb[tab]
    rows=[[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]
    hdr=[txt(x) for x in rows[0]]
    cur=m.get('month',''); prior=m.get('prior_month',''); py=m.get('prior_year_month','')
    cfy=m.get('cur_fy',''); pfy=m.get('prior_fy','')
    ROLE={f'{prior} Act':'prior_act', f'{cur} Act':'cur_act', f'{cur} Bud':'cur_bud',
          'M-o-M Bud Var (%)':'mom_bud_var', f'YTD {cfy} Act':'ytd_act', f'YTD {cfy} Bud':'ytd_bud',
          'Y-o-Y Bud Var (%)':'yoy_bud_var', f'YTD {pfy} Act':'pfy_ytd_act', 'Y-o-Y Var (%)':'yoy_var',
          f'{py} Act':'py_month_act', 'YoY month Var(%)':'yoy_month_var'}
    cm={ROLE[h]:j for j,h in enumerate(hdr) if h in ROLE}
    out=[]; bu=None
    for row in rows[1:]:
        label=txt(row[0]); has_nums=any(isinstance(x,(int,float)) for x in row[1:])
        if label in BUS and not has_nums: bu=label; continue
        if bu and label and has_nums:
            rec={'bu':bu,'metric':label}
            for role,j in cm.items(): rec[role]=num(row[j]) if j<len(row) else None
            out.append(rec)
    return out, m

def validate(data):
    """returns (ok, messages). Checks BU coverage + EBITDA = Revenue - Cost (current)."""
    msg=[]; ok=True
    bus=sorted({r['bu'] for r in data})
    if bus!=sorted(BUS): ok=False; msg.append(f"BU set {bus} != {sorted(BUS)}")
    idx={(r['bu'],r['metric']):r for r in data}
    for bu in ['DaaS','Distribution','Martech']:
        rev=idx.get((bu,'Revenue'),{}).get('cur_act'); cost=idx.get((bu,'Cost'),{}).get('cur_act'); eb=idx.get((bu,'EBITDA'),{}).get('cur_act')
        if None in (rev,cost,eb): ok=False; msg.append(f"{bu}: missing Revenue/Cost/EBITDA"); continue
        if abs((rev-cost)-eb)>1.0: ok=False; msg.append(f"{bu}: EBITDA {eb} != Rev-Cost {round(rev-cost,1)}")
    # Consolidated uses Net revenue
    gr=idx.get(('Consolidated','Gross Revenue'),{}).get('cur_act'); cc=idx.get(('Consolidated','Cost'),{}).get('cur_act'); ce=idx.get(('Consolidated','EBITDA'),{}).get('cur_act')
    if None not in (gr,cc,ce) and abs((gr-cc)-ce)>1.0: ok=False; msg.append(f"Consolidated: EBITDA {ce} != GrossRev-Cost {round(gr-cc,1)}")
    return ok, (msg or ["all checks passed"])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/uploads/Consolidated_P_L_Summary_Jun_26.xlsx'
    data,meta=convert(path); ok,msg=validate(data)
    print("META:",{k:meta.get(k) for k in ('month','cur_fy','prior_fy','prior_month','prior_year_month')})
    print("rows:",len(data),"| BUs:",sorted({r['bu'] for r in data}))
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg))
    print("\n--- DaaS (all metrics, incl %) ---")
    for r in [x for x in data if x['bu']=='DaaS']: print(json.dumps(r,ensure_ascii=False))
    json.dump(data,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/consol_pnl_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
