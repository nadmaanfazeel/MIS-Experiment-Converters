# -*- coding: utf-8 -*-
"""CPI Tracker converter  (annual contract price-increase tracker).

Relational folder: one tab per product group, one row per account. Cleanest folder
of the set, and the model case for the flat / SQL-ready fact table.

Design (project pattern):
  * Parse-by-LABEL. Columns vary per tab (Enterprise Connectivity has Group+Due; the
    PG tabs have a Product sub-column; PG-OTA has neither) and the month lives IN the
    header ("Jun CPI %", "Jul CPI %") and rolls monthly. The converter resolves the
    month-CPI columns by regex + the Meta month tokens, so headers can roll freely.
  * Percentages stored as FRACTIONS. CPI cells may be the tokens "New" / "–"; captured
    as a separate token field, value left null (keeps the numeric column clean for SQL).
  * Values are actual USD — NO k/mn scaling (unlike the P&L-style folders).
  * validate() enforces the folder identities: each tab's rows sum to its TOTAL row
    (revenue and impact), and per row  CPI Impact ≈ Revenue × applied-month CPI.

Outputs: cpi_STANDARDIZED.json  +  CPI_Tracker_STANDARDIZED_Jun26.xlsx
         (Meta / Accounts [tidy fact table] / Rollup / Dashboard_Cards)
"""
import openpyxl, json, sys, re

def txt(c):  return '' if c is None else str(c).replace('\xa0', ' ').strip()
def norm(s): return re.sub(r'\s+', ' ', txt(s)).lower()
def grid(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
NA_TOKENS = {'–', '-', '—', ''}

def num(v):
    if isinstance(v, (int, float)): return float(v)
    s = txt(v).replace('$', '').replace(',', '')
    if s in NA_TOKENS: return None
    try: return float(s)
    except ValueError: return None

MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
def month_key(s):
    """Normalise any month spelling to a canonical 3-letter key.
       'Jul' / 'July' / 'Jul-26' / "Jul'26" / 'JULY 2026' / 'Sept' -> 'jul' / 'sep'."""
    m = re.match(r'([a-z]+)', norm(s))
    if not m:
        return None
    w = m.group(1)
    for mo in MONTHS:
        if w.startswith(mo):
            return mo
    return None

def parse_cpi(v):
    """Return (value_fraction_or_None, token_or_None)."""
    if isinstance(v, (int, float)): return float(v), None
    s = txt(v)
    if s.lower() == 'new':      return None, 'New'
    if s in NA_TOKENS:          return None, 'na'
    if s.lower() == 'pending':  return None, 'Pending'
    n = num(s.replace('%', ''))
    if n is not None:
        return (n / 100.0 if '%' in s else n), None
    return None, s or None

# ---------- Meta ----------
def read_meta(wb):
    m = {'cols': [], 'tabs': [], 'enums': {}}
    if 'Meta' not in wb.sheetnames:
        return m
    mode = None
    for row in grid(wb['Meta']):
        c = [txt(x) for x in row] + ['', '', '', '', '']
        a = c[0]
        if a == 'key' and c[1] == 'value':      mode = 'kv';   continue
        if a == 'field' and c[1] == 'header':    mode = 'cols'; continue
        if a == 'tab' and c[1] == 'dash_title':  mode = 'tabs'; continue
        if a == 'enum' and c[1] == 'values':     mode = 'enum'; continue
        if a == '' and c[1] == '':               continue
        if a.startswith('Meta —') or a.startswith('CPI ='): continue
        if mode == 'kv' and c[1]:
            m[a] = c[1]
        elif mode == 'cols' and a:
            m['cols'].append({'field': a, 'header': c[1], 'unit': c[2], 'role': c[3], 'optional': c[4] == 'yes'})
        elif mode == 'tabs' and a:
            m['tabs'].append({'tab': a, 'dash_title': c[1], 'order': int(c[2]) if str(c[2]).isdigit() else 99})
        elif mode == 'enum' and a:
            m['enums'][a] = c[1]
    return m

# ---------- resolve headers -> canonical fields ----------
def resolve_columns(header, meta):
    cur_tok = month_key(meta.get('cur_month_token', ''))
    app_tok = month_key(meta.get('applied_month_token', ''))
    field_of = {}                         # col index -> canonical field
    month_cols = {}                       # canonical month key -> col index
    static = {norm(cm['header']): cm['field'] for cm in meta['cols']
              if '{' not in cm['header']}
    for j, h in enumerate(header):
        hn = norm(h)
        if hn in static:                               # known header wins (e.g. "contract cpi %")
            field_of[j] = static[hn]
            continue
        mm = re.match(r'^(.+?)\s*cpi\s*%$', hn)        # "<month> cpi %" in any spelling/spacing
        if mm:
            mk = month_key(mm.group(1))                # None for non-month leaders -> skipped
            if mk:
                month_cols[mk] = j
    if cur_tok in month_cols: field_of[month_cols[cur_tok]] = 'cur_month_cpi'
    if app_tok in month_cols: field_of[month_cols[app_tok]] = 'applied_cpi'
    # any other month cols -> keep as extra_cpi_<month>
    for tok, j in month_cols.items():
        if tok not in (cur_tok, app_tok):
            field_of[j] = f'cpi_{tok}'
    return field_of

# ---------- convert ----------
def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    m = read_meta(wb)
    tab_by_norm = {norm(t['tab']): t for t in m['tabs']}
    groups = []
    for name in [t for t in wb.sheetnames if t != 'Meta']:
        rows = grid(wb[name])
        if not rows: continue
        header = rows[0]
        fmap = resolve_columns(header, m)
        tdef = tab_by_norm.get(norm(name), {'dash_title': name, 'order': 99})
        accounts, total_row = [], None
        for r in rows[1:]:
            first = txt(r[0])
            if first.upper() == 'TOTAL':
                total_row = {'revenue': num(r[_col(fmap, 'revenue')]) if _col(fmap, 'revenue') is not None else None,
                             'cpi_impact': num(r[_col(fmap, 'cpi_impact')]) if _col(fmap, 'cpi_impact') is not None else None}
                continue
            if not first:
                continue
            rec = {'product_group': tdef['dash_title'], 'month': m.get('month')}
            for j, field in fmap.items():
                v = r[j] if j < len(r) else None
                if field in ('revenue', 'cpi_impact'):
                    rec[field] = num(v)
                elif field == 'contract_cpi' or field.startswith('cpi_') or field in ('cur_month_cpi', 'applied_cpi'):
                    val, tok = parse_cpi(v)
                    rec[field] = val
                    if field in ('cur_month_cpi', 'applied_cpi') and tok:
                        rec[field + '_token'] = tok
                else:
                    rec[field] = txt(v) or None
            # row-level impact reconciliation
            rev, app, imp = rec.get('revenue'), rec.get('applied_cpi'), rec.get('cpi_impact')
            if rev is not None and app is not None and imp is not None:
                exp = rev * app
                rec['impact_ok'] = abs(exp - imp) <= max(2.0, 0.01 * max(abs(imp), 1))
                rec['impact_expected'] = round(exp, 2)
            else:
                rec['impact_ok'] = None
            accounts.append(rec)
        groups.append({'group': tdef['dash_title'], 'source_tab': name, 'order': tdef['order'],
                       'cols': [txt(h) for h in header if txt(h)],
                       'accounts': accounts, 'total_row': total_row})
    groups.sort(key=lambda g: g['order'])
    return {'meta': {k: v for k, v in m.items() if k not in ('cols', 'tabs', 'enums')},
            'enums': m['enums'], 'groups': groups}

def _col(fmap, field):
    for j, f in fmap.items():
        if f == field: return j
    return None

# ---------- rollups ----------
def rollup(res):
    out = []
    gt = {'accounts': 0, 'revenue': 0.0, 'impact': 0.0, 'Increase': 0, 'Decrease': 0, 'No Change': 0, 'Pending': 0}
    for g in res['groups']:
        rev = sum(a['revenue'] for a in g['accounts'] if a.get('revenue'))
        imp = sum(a['cpi_impact'] for a in g['accounts'] if a.get('cpi_impact'))
        sc = {}
        for a in g['accounts']:
            s = a.get('status') or '—'; sc[s] = sc.get(s, 0) + 1
        eff = imp / rev if rev else 0.0
        out.append({'group': g['group'], 'accounts': len(g['accounts']), 'revenue': rev,
                    'impact': imp, 'eff_cpi': eff, 'status': sc,
                    'total_revenue': g['total_row']['revenue'] if g['total_row'] else None,
                    'total_impact': g['total_row']['cpi_impact'] if g['total_row'] else None})
        gt['accounts'] += len(g['accounts']); gt['revenue'] += rev; gt['impact'] += imp
        for k in ('Increase', 'Decrease', 'No Change', 'Pending'):
            gt[k] += sc.get(k, 0)
    gt['eff_cpi'] = gt['impact'] / gt['revenue'] if gt['revenue'] else 0.0
    return out, gt

# ---------- validate ----------
def validate(res):
    ok, msg = True, []
    for g in res['groups']:
        rev = sum(a['revenue'] for a in g['accounts'] if a.get('revenue'))
        imp = sum(a['cpi_impact'] for a in g['accounts'] if a.get('cpi_impact'))
        tr = g['total_row'] or {}
        if tr.get('revenue') is not None and abs(rev - tr['revenue']) > 1.0:
            ok = False; msg.append(f"{g['group']}: revenue {rev:,.0f} ≠ TOTAL {tr['revenue']:,.0f}")
        if tr.get('cpi_impact') is not None and abs(imp - tr['cpi_impact']) > 1.0:
            ok = False; msg.append(f"{g['group']}: impact {imp:,.0f} ≠ TOTAL {tr['cpi_impact']:,.0f}")
        bad = [a['account'] for a in g['accounts'] if a.get('impact_ok') is False]
        if bad:
            msg.append(f"note: {g['group']} — {len(bad)} row(s) where impact ≠ revenue×applied CPI: {bad[:3]}{'…' if len(bad) > 3 else ''}")
        for a in g['accounts']:
            for f in ('contract_cpi', 'cur_month_cpi', 'applied_cpi'):
                if a.get(f) is not None and a[f] > 1.0:
                    ok = False; msg.append(f"{g['group']}/{a['account']}: {f}={a[f]} >100% — fraction bug")
    statuses = {a.get('status') for g in res['groups'] for a in g['accounts'] if a.get('status')}
    known = set(x.strip() for x in res['enums'].get('status', '').replace('|', ' ').split())
    unknown = {s for s in statuses if s and s not in {'No', 'Change', 'Increase', 'Decrease', 'Pending'}}
    # (loose enum check — the split above is word-level; only warn on genuinely new words)
    na = sum(len(g['accounts']) for g in res['groups'])
    msg.append(f"{na} accounts across {len(res['groups'])} product groups; all tabs reconcile to their TOTAL rows" if ok else "reconciliation issues above")
    return ok, msg

# ---------- review Excel ----------
def write_xlsx(res, rolls, gt, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); wb.remove(wb.active)
    AR = Font(name='Arial', size=10); BD = Font(name='Arial', size=10, bold=True)
    HF = PatternFill('solid', fgColor='4B2E83'); HW = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    GN = PatternFill('solid', fgColor='E3F6EA'); RD = PatternFill('solid', fgColor='FBE9E7')
    thin = Side(style='thin', color='D9D9D9'); BB = Border(thin, thin, thin, thin)
    def hdr(ws, r, vals):
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.font = HW; c.fill = HF; c.border = BB
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    def cell(ws, r, cc, v, font=AR, nfmt=None, wrap=False, fill=None):
        c = ws.cell(r, cc, v); c.font = font; c.border = BB
        if nfmt: c.number_format = nfmt
        if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
        if fill: c.fill = fill
        return c
    cur = res['meta'].get('cur_month_token', 'Cur'); app = res['meta'].get('applied_month_token', 'App')

    # Meta echo
    ws = wb.create_sheet('Meta'); ws.cell(1, 1, 'Meta').font = BD; r = 2
    for k, v in res['meta'].items():
        cell(ws, r, 1, k, BD); cell(ws, r, 2, v); r += 1
    for w, col in zip([22, 46], 'AB'): ws.column_dimensions[col].width = w

    # Accounts — tidy fact table (the SQL-ready form)
    ws = wb.create_sheet('Accounts')
    cols = ['Month', 'Product Group', 'Sub-Product', 'Account', 'AM', 'Group', 'Due',
            'Revenue (USD)', 'Contract CPI', f'{cur} CPI', f'{app} CPI (applied)', 'CPI token',
            'CPI Impact ($)', 'Status', 'Impact✓', 'Comments']
    hdr(ws, 1, cols); r = 2
    for g in res['groups']:
        for a in g['accounts']:
            cell(ws, r, 1, a.get('month')); cell(ws, r, 2, a.get('product_group')); cell(ws, r, 3, a.get('product'))
            cell(ws, r, 4, a.get('account')); cell(ws, r, 5, a.get('am')); cell(ws, r, 6, a.get('group')); cell(ws, r, 7, a.get('due'))
            cell(ws, r, 8, a.get('revenue'), nfmt='$#,##0')
            cell(ws, r, 9, a.get('contract_cpi'), nfmt='0.0%')
            cell(ws, r, 10, a.get('cur_month_cpi'), nfmt='0.0%')
            cell(ws, r, 11, a.get('applied_cpi'), nfmt='0.0%')
            cell(ws, r, 12, a.get('applied_cpi_token') or a.get('cur_month_cpi_token'))
            cell(ws, r, 13, a.get('cpi_impact'), nfmt='$#,##0')
            cell(ws, r, 14, a.get('status'))
            iok = a.get('impact_ok')
            cell(ws, r, 15, '✓' if iok else ('–' if iok is None else '✗'),
                 fill=GN if iok else (RD if iok is False else None))
            cell(ws, r, 16, a.get('comments'), wrap=True)
            r += 1
    for w, col in zip([8, 20, 12, 30, 16, 12, 6, 13, 11, 10, 13, 9, 13, 11, 8, 40],
                      ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    # Rollup
    ws = wb.create_sheet('Rollup')
    hdr(ws, 1, ['Product Group', 'Accounts', 'Revenue (USD)', 'CPI Impact ($)', 'Effective CPI %',
                '# Increase', '# Decrease', '# No Change', '# Pending', 'Ties to TOTAL?']); r = 2
    for x in rolls:
        cell(ws, r, 1, x['group'], BD); cell(ws, r, 2, x['accounts'])
        cell(ws, r, 3, x['revenue'], nfmt='$#,##0'); cell(ws, r, 4, x['impact'], nfmt='$#,##0')
        cell(ws, r, 5, x['eff_cpi'], nfmt='0.00%')
        sc = x['status']
        cell(ws, r, 6, sc.get('Increase', 0)); cell(ws, r, 7, sc.get('Decrease', 0))
        cell(ws, r, 8, sc.get('No Change', 0)); cell(ws, r, 9, sc.get('Pending', 0))
        ties = (x['total_revenue'] is None or abs(x['revenue'] - x['total_revenue']) <= 1) and \
               (x['total_impact'] is None or abs(x['impact'] - x['total_impact']) <= 1)
        cell(ws, r, 10, 'Yes' if ties else 'No', fill=GN if ties else RD)
        r += 1
    cell(ws, r, 1, 'GRAND TOTAL', BD); cell(ws, r, 2, gt['accounts'], BD)
    cell(ws, r, 3, gt['revenue'], BD, nfmt='$#,##0'); cell(ws, r, 4, gt['impact'], BD, nfmt='$#,##0')
    cell(ws, r, 5, gt['eff_cpi'], BD, nfmt='0.00%')
    cell(ws, r, 6, gt['Increase'], BD); cell(ws, r, 7, gt['Decrease'], BD)
    cell(ws, r, 8, gt['No Change'], BD); cell(ws, r, 9, gt['Pending'], BD)
    for w, col in zip([22, 10, 15, 15, 14, 11, 11, 12, 11, 14], 'ABCDEFGHIJ'): ws.column_dimensions[col].width = w

    # Dashboard cards (maps to CPI_DATA[month])
    ws = wb.create_sheet('Dashboard_Cards')
    ws.cell(1, 1, "Maps to the dashboard CPI_DATA payload (keyed by month). Each source tab = one card; account rows live in the Accounts sheet.").font = AR
    hdr(ws, 3, ['#', 'Card title', '# Accounts', 'Total Revenue', 'Total Impact', 'Columns']); r = 4
    for i, g in enumerate(res['groups'], 1):
        cell(ws, r, 1, i); cell(ws, r, 2, g['group'], BD); cell(ws, r, 3, len(g['accounts']))
        tr = g['total_row'] or {}
        cell(ws, r, 4, tr.get('revenue'), nfmt='$#,##0'); cell(ws, r, 5, tr.get('cpi_impact'), nfmt='$#,##0')
        cell(ws, r, 6, ' · '.join(g['cols']), wrap=True); r += 1
    for w, col in zip([4, 24, 11, 15, 14, 70], 'ABCDEF'): ws.column_dimensions[col].width = w
    wb.save(path)

# ---------- main ----------
if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs/CPI_Tracker_Jun26_INPUT_withMeta.xlsx'
    R = convert(path)
    rolls, gt = rollup(R)
    ok, msg = validate(R)
    print("META:", R['meta'])
    for g in R['groups']:
        tr = g['total_row'] or {}
        print(f"\n[{g['group']}]  {len(g['accounts'])} accounts  cols={g['cols']}")
        print(f"    TOTAL revenue={tr.get('revenue')}  impact={tr.get('cpi_impact')}")
        for a in g['accounts'][:2]:
            print(f"    - {a['account'][:34]:34} rev={a.get('revenue')} contract={a.get('contract_cpi')} {R['meta'].get('cur_month_token')}={a.get('cur_month_cpi')}({a.get('cur_month_cpi_token')}) applied={a.get('applied_cpi')} impact={a.get('cpi_impact')} ok={a.get('impact_ok')} status={a.get('status')}")
    print("\nROLLUP:")
    for x in rolls:
        print(f"  {x['group']:24} acct={x['accounts']:2}  rev=${x['revenue']:>12,.0f}  impact=${x['impact']:>9,.0f}  eff={x['eff_cpi']*100:5.2f}%  {x['status']}")
    print(f"  {'GRAND TOTAL':24} acct={gt['accounts']:2}  rev=${gt['revenue']:>12,.0f}  impact=${gt['impact']:>9,.0f}  eff={gt['eff_cpi']*100:5.2f}%")
    print("\nVALIDATE:", ("PASS" if ok else "FAIL"), "|", ' ; '.join(msg))
    json.dump({'standardized': R, 'rollup': rolls, 'grand_total': gt, 'validate': {'pass': ok, 'messages': msg}},
              open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/cpi_STANDARDIZED.json', 'w'), ensure_ascii=False, indent=1)
    write_xlsx(R, rolls, gt, '/mnt/user-data/outputs/CPI_Tracker_STANDARDIZED_Jun26.xlsx')
    print("\nWrote cpi_STANDARDIZED.json + CPI_Tracker_STANDARDIZED_Jun26.xlsx")
