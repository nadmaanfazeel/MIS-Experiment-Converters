# -*- coding: utf-8 -*-
"""Revenue by BU converter. Single deep hierarchy (BU -> sub-BU -> leaf) with
INCONSISTENT indentation, so the tree is defined by a Meta node-map (node->role:
bu/sub_bu/total/adjustment; anything else = leaf). Values in $'000; YoY as fraction."""
import openpyxl, json, sys

def num(c): return round(float(c),3) if isinstance(c,(int,float)) else None
def txt(c): return '' if c is None else str(c).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

def read_meta(wb):
    m={'nodes':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='node' and b=='role': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['nodes'][a]=b
    return m

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    ws=next(wb[t] for t in wb.sheetnames if t!='Meta'); rows=grid(ws)
    hi=next(i for i,r in enumerate(rows) if any("BU" in txt(x) and "USD" in txt(x) for x in r))
    H=[txt(x) for x in rows[hi]]; lc=next(j for j,h in enumerate(H) if 'BU' in h)
    # value columns
    c_cur=next(j for j,h in enumerate(H) if 'FY27' in h.replace(' ',''))
    c_pfy=next(j for j,h in enumerate(H) if 'FY26' in h.replace(' ',''))
    c_yoy=next((j for j,h in enumerate(H) if 'Growth' in h or 'YoY' in h or 'Y-o-Y' in h),None)
    out=[]; cur_bu=None; cur_sub=None
    for row in rows[hi+1:]:
        node=txt(row[lc]) if lc<len(row) else ''
        if not node: continue
        role=m['nodes'].get(node,'leaf')
        if role=='bu': cur_bu=node; cur_sub=None; level=0; parent=None
        elif role=='sub_bu': cur_sub=node; level=1; parent=cur_bu
        elif role in ('total','adjustment'): level=0; parent=None; cur_bu=None; cur_sub=None
        else: level=2 if cur_sub else 1; parent=cur_sub or cur_bu; role='leaf'
        out.append({'node':node,'role':role,'level':level,'parent':parent,'bu':cur_bu,
                    'ytd_cur':num(row[c_cur]) if c_cur<len(row) else None,
                    'ytd_pfy':num(row[c_pfy]) if c_pfy<len(row) else None,
                    'yoy_growth':num(row[c_yoy]) if c_yoy is not None and c_yoy<len(row) else None})
    return {'meta':{k:v for k,v in m.items() if k!='nodes'}, 'rows':out}

def validate(res):
    ok=True; msg=[]; rows=res['rows']; idx={r['node']:r for r in rows}
    def kids(parent): return [r for r in rows if r['parent']==parent]
    for r in rows:
        if r['role'] in ('bu','sub_bu'):
            ch=kids(r['node'])
            if ch:
                s=sum(c['ytd_cur'] or 0 for c in ch)
                if abs(s-(r['ytd_cur'] or 0))>max(1.0,0.01*abs(r['ytd_cur'] or 1)):
                    ok=False; msg.append(f"{r['node']} {round(r['ytd_cur'],1)} != sum(children) {round(s,1)}")
    # Net == sum(bu); Gross == Net + Ad spend
    bus=[r for r in rows if r['role']=='bu']
    net=next((r for r in rows if 'Net Revenue' in r['node']),None)
    if net and bus:
        s=sum(b['ytd_cur'] or 0 for b in bus)
        if abs(s-(net['ytd_cur'] or 0))>max(1.0,0.01*abs(net['ytd_cur'] or 1)): ok=False; msg.append(f"Net {round(net['ytd_cur'],1)} != sum(BUs) {round(s,1)}")
    gross=next((r for r in rows if 'Gross Revenue' in r['node']),None)
    ad=next((r for r in rows if r['role']=='adjustment'),None)
    if gross and net and ad:
        if abs(((net['ytd_cur'] or 0)+(ad['ytd_cur'] or 0))-(gross['ytd_cur'] or 0))>1.0: ok=False; msg.append("Gross != Net + Ad spend")
    return ok,(msg or ['hierarchy sums + Net/Gross identities hold'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/Revenue_by_BU_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'])
    print("rows:",len(R['rows']),"| BUs:",[r['node'] for r in R['rows'] if r['role']=='bu'],"| sub_bus:",[r['node'] for r in R['rows'] if r['role']=='sub_bu'])
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:6]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/revbu_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
