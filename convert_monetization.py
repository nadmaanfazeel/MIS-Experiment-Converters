# -*- coding: utf-8 -*-
"""Monetization folder converter (10 tabs).
Reads a Meta tab: temporal keys + a product map (tab->product,bu,schema).
Extracts, per product: annual/quarterly COHORT table, MONTHLY chart data, and DEAL rows (if any).
Consolidated: cohort + orderbook trend + orderbook-by-BU. Roles fixed; #DIV/0! -> null."""
import openpyxl, datetime, re, json, sys

def num(c):
    if isinstance(c,(int,float)): return round(float(c),3)
    return None
def txt(c): return '' if c is None else (c.strftime('%b-%y') if isinstance(c,datetime.datetime) else str(c)).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]

def parse_month(x):
    t=(txt(x) or '').replace("'","-").replace("’","-").replace(" ","-")
    t=re.sub(r'-+','-',t)
    try: return datetime.datetime.strptime(t,"%b-%y")
    except: return None
def month_window(cur_label, trailing):
    end=parse_month(cur_label)
    if not end: return None,None,None,None
    et=end.year*12+(end.month-1); st=et-(int(trailing)-1)
    return st,et,("%d-%02d"%(st//12,st%12+1)),("%d-%02d"%(et//12,et%12+1))
def _lbl(idx):
    import calendar
    y=idx//12; mo=idx%12+1
    return "%s'%02d"%(calendar.month_abbr[mo], y%100)

def read_meta(wb):
    m={'products':[]}
    if 'Meta' not in wb.sheetnames: return m
    ws=wb['Meta']; rows=grid(ws); mode='kv'
    for row in rows:
        a=txt(row[0]) if len(row)>0 else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='tab' and b=='product': mode='map'; continue
        if mode=='kv' and a.lower()!='key':
            m[a]=b
        elif mode=='map':
            m['products'].append({'tab':a,'product':b,
                                  'bu':txt(row[2]) if len(row)>2 else '',
                                  'schema':txt(row[3]) if len(row)>3 else 'annual'})
    return m

METRICS=['Opp Value','Total Invoicing','Orderbook','Monetization %']
def find_cohort_header(rows):
    for i,row in enumerate(rows):
        cells=[txt(x) for x in row]
        if 'Opp Value' in cells and 'Orderbook' in cells and 'Monetization %' in cells:
            oc=cells.index('Opp Value')
            return i, oc, oc-1   # header row idx, opp col, label col
    return None,None,None

def read_cohort(rows, hi, oc, lc):
    out=[]
    for row in rows[hi+1:]:
        period=txt(row[lc]) if lc<len(row) else ''
        if period=='' : 
            # stop if we hit a fully blank label after we've started
            if out: break
            else: continue
        vals=[row[oc+k] if oc+k<len(row) else None for k in range(4)]
        if not any(isinstance(v,(int,float)) for v in vals): 
            if out: break
            else: continue
        out.append({'period':period,
                    'opp_value':num(vals[0]),'total_invoicing':num(vals[1]),
                    'orderbook':num(vals[2]),'monetization_pct':num(vals[3])})
        if period.lower()=='total': break
    return out

def find_monthly(rows):
    for i,row in enumerate(rows):
        cells=[txt(x) for x in row]
        if 'Month' in cells and any('Opp Value' in c for c in cells) and any('Monetised' in c for c in cells):
            mc=cells.index('Month'); return i, mc
    return None,None
def read_monthly(rows, hi, mc):
    out=[]
    for row in rows[hi+1:]:
        mon=txt(row[mc]) if mc<len(row) else ''
        if not re.match(r"^[A-Za-z]{3}['-]\d{2}$",mon):
            if out: break
            else: continue
        out.append({'month':mon,'opp_value':num(row[mc+1]),'monetised':num(row[mc+2]),'orderbook':num(row[mc+3])})
    return out

def find_deals(rows):
    for i,row in enumerate(rows):
        cells=[txt(x) for x in row]
        if 'Opp Name' in cells and 'Orderbook' in cells and 'Status' in cells:
            return i,[txt(x) for x in row]
    return None,None
def read_deals(rows,hi,header):
    idx={h:j for j,h in enumerate(header)}
    def g(row,col):
        j=idx.get(col); return row[j] if (j is not None and j<len(row)) else None
    out=[]
    for row in rows[hi+1:]:
        name=txt(g(row,'Opp Name'))
        if not name:
            if out: break
            continue
        rec={'opp_name':name,'opp_value':num(g(row,'Opp Value')),'product':txt(g(row,'Product')),
             'closed_fy':txt(g(row,'Opp Closed FY')),'close_month':txt(g(row,'Close Month')),
             'total_monetised':num(g(row,'Total Monetised')),'monetization_pct':num(g(row,'Monetization %')),
             'orderbook':num(g(row,'Orderbook')),'status':txt(g(row,'Status'))}
        if name.lower().startswith('total'): rec['total']=True
        out.append(rec)
        if rec.get('total'): break
    return out

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    tm=int(m.get('trailing_months',12) or 12)
    st,et,_,_=month_window(m.get('month',''),tm)
    def in_window(monlabel):
        dt=parse_month(monlabel)
        if not dt or st is None: return True
        x=dt.year*12+(dt.month-1); return st<=x<=et
    pmap={p['tab']:p for p in m['products']}
    res={'meta':{k:v for k,v in m.items() if k!='products'}, 'consolidated':{}, 'products':[]}
    if st is not None:
        res['meta']['window_start']=_lbl(st); res['meta']['window_end']=_lbl(et); res['meta']['trailing_months']=tm
    for tab in wb.sheetnames:
        if tab in ('Meta',): continue
        ws=wb[tab]; rows=grid(ws)
        hi,oc,lc=find_cohort_header(rows)
        if tab=='Consolidated':
            res['consolidated']['cohorts']=read_cohort(rows,hi,oc,lc) if hi is not None else []
            # orderbook trend (Month|Orderbook) + BU chart (BU|Orderbook) live in far-right cols
            trend=[]; bu=[]
            for row in rows:
                cells=[txt(x) for x in row]
                for j,c in enumerate(cells):
                    if re.match(r"^[A-Za-z]{3}'\d{2}$",c) and j+1<len(row) and isinstance(row[j+1],(int,float)):
                        trend.append({'month':c,'orderbook':num(row[j+1])})
                # BU rows: label like 'OTA','EC','DaaS - HospiBI' followed by a number in the BU-chart area
            # BU chart: find 'BU' header then read
            for i,row in enumerate(rows):
                if 'BU' in [txt(x) for x in row] and 'Orderbook' in [txt(x) for x in row]:
                    bc=[txt(x) for x in row].index('BU')
                    for r2 in rows[i+1:]:
                        lbl=txt(r2[bc]) if bc<len(r2) else ''
                        if lbl and isinstance(r2[bc+1] if bc+1<len(r2) else None,(int,float)):
                            bu.append({'bu':lbl,'orderbook':num(r2[bc+1])})
                        elif lbl=='' : break
                    break
            res['consolidated']['orderbook_trend']=[t for t in trend if in_window(t['month'])]
            res['consolidated']['orderbook_by_bu']=bu
            continue
        info=pmap.get(tab,{'product':tab,'bu':'','schema':'annual'})
        rec={'tab':tab,'product':info['product'],'bu':info['bu'],'schema':info['schema'],
             'cohorts':read_cohort(rows,hi,oc,lc) if hi is not None else []}
        mi,mc=find_monthly(rows)
        if mi is not None: rec['monthly']=[mo for mo in read_monthly(rows,mi,mc) if in_window(mo['month'])]
        di,dh=find_deals(rows)
        if di is not None: rec['deals']=read_deals(rows,di,dh)
        res['products'].append(rec)
    return res

def validate(res):
    msgs=[]; ok=True
    # identity: opp ~= invoicing + orderbook  (per cohort, where all present)
    def check(rowset,label):
        nonlocal ok
        for c in rowset:
            o,i,b=c.get('opp_value'),c.get('total_invoicing'),c.get('orderbook')
            if None not in (o,i,b) and o>1 and abs(o-(i+b))>max(1.0,0.02*o):
                ok=False; msgs.append(f"{label}/{c['period']}: opp {o} != inv+ob {round(i+b,1)}")
    check(res['consolidated'].get('cohorts',[]),'Consolidated')
    for p in res['products']: check(p['cohorts'],p['product'])
    exp={p['tab'] for p in res['meta'].get('products',[])} if False else None
    return ok,(msgs or ['all identity checks passed'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/uploads/Monetization_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msgs=validate(R)
    print("META:",R['meta'])
    print("products:",[ (p['product'],p['bu'],p['schema'],len(p['cohorts']),'mon:'+str(len(p.get('monthly',[])))) for p in R['products']])
    print("consolidated cohorts:",len(R['consolidated'].get('cohorts',[])),"| trend pts:",len(R['consolidated'].get('orderbook_trend',[])),"| by_bu:",len(R['consolidated'].get('orderbook_by_bu',[])))
    print("VALIDATE:",("PASS" if ok else "WARN"),"|",'; '.join(msgs[:5]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/monetization_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
