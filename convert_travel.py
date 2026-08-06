# -*- coding: utf-8 -*-
"""Travel Expense converter. Meta-driven.
- 'leader_table' tab: Leader | Approved | YTD Bud | [Mmm'YY Expense ...FY-to-date] | YTD Expense | Variance | Comments
  The monthly columns run Apr->current month of the FY and GROW each month (read dynamically).
- 'sojern_dept' tab: Dept x (Property/Corporate/Destination/Total) matrix.
Values are FULL USD (not '000)."""
import openpyxl, datetime, re, json, sys

def num(c): return round(float(c),2) if isinstance(c,(int,float)) else None
def txt(c): return '' if c is None else (c.strftime("%b'%y") if isinstance(c,datetime.datetime) else str(c)).strip()
def grid(ws): return [[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]
MON_EXP=re.compile(r"^([A-Za-z]{3})'?(\d{2}) Expense$")

def read_meta(wb):
    m={'tabs':{}}
    if 'Meta' not in wb.sheetnames: return m
    mode='kv'
    for row in grid(wb['Meta']):
        a=txt(row[0]) if row else ''; b=txt(row[1]) if len(row)>1 else ''
        if a=='' and b=='': continue
        if a=='tab' and b=='type': mode='map'; continue
        if mode=='kv' and a.lower()!='key': m[a]=b
        elif mode=='map': m['tabs'][a]=b
    return m

def conv_leader(ws):
    rows=grid(ws)
    hi=next((i for i,r in enumerate(rows) if 'Leader' in [txt(x) for x in r] and 'Variance' in [txt(x) for x in r]),None)
    if hi is None: return None
    H=[txt(x) for x in rows[hi]]
    def col(pred):
        for j,h in enumerate(H):
            if pred(h): return j
        return None
    c_leader=col(lambda h:h=='Leader'); c_appr=col(lambda h:h=='Approved')
    c_ybud=col(lambda h:'YTD Bud' in h); c_yexp=col(lambda h:'YTD Expense' in h)
    c_var=col(lambda h:h=='Variance'); c_com=col(lambda h:'Comment' in h)
    mcols=[(j,txt(H[j])) for j,h in enumerate(H) if MON_EXP.match(h)]
    months=[MON_EXP.match(h).group(1)+"'"+MON_EXP.match(h).group(2) for _,h in mcols]
    out=[]
    for row in rows[hi+1:]:
        leader=txt(row[c_leader]) if c_leader<len(row) else ''
        if not leader: 
            if out: break
            continue
        rec={'leader':leader,'approved':num(row[c_appr]) if c_appr is not None and c_appr<len(row) else None,
             'ytd_bud':num(row[c_ybud]) if c_ybud is not None and c_ybud<len(row) else None,
             'monthly':{},'ytd_expense':num(row[c_yexp]) if c_yexp is not None and c_yexp<len(row) else None,
             'variance':num(row[c_var]) if c_var is not None and c_var<len(row) else None,
             'comments':txt(row[c_com]) if c_com is not None and c_com<len(row) else ''}
        for (j,_),mlabel in zip(mcols,months): rec['monthly'][mlabel]=num(row[j]) if j<len(row) else None
        if leader.lower()=='total': rec['total']=True
        out.append(rec)
        if rec.get('total'): break
    return {'months':months,'rows':out}

def conv_sojern(ws):
    rows=grid(ws)
    hi=next((i for i,r in enumerate(rows) if 'Dept' in [txt(x) for x in r] and 'Property' in [txt(x) for x in r]),None)
    if hi is None: return None
    H=[txt(x) for x in rows[hi]]; dc=H.index('Dept'); ents=[H[j] for j in range(dc+1,len(H)) if H[j]]
    out=[]
    for row in rows[hi+1:]:
        dept=txt(row[dc]) if dc<len(row) else ''
        if not dept: continue
        rec={'dept':dept}
        for j in range(dc+1,len(H)):
            if H[j]: rec[H[j]]=num(row[j]) if j<len(row) else None
        out.append(rec)
    return {'entities':ents,'rows':out}

def convert(path):
    wb=openpyxl.load_workbook(path,data_only=True); m=read_meta(wb)
    res={'meta':{k:v for k,v in m.items() if k!='tabs'}}
    for tab in wb.sheetnames:
        if tab=='Meta': continue
        typ=m['tabs'].get(tab,'leader_table' if 'only Sojern' not in tab else 'sojern_dept')
        if typ=='sojern_dept': res['sojern_dept']=conv_sojern(wb[tab])
        else: res['leader_table']=conv_leader(wb[tab])
    return res

def validate(res):
    ok=True; msg=[]
    lt=res.get('leader_table',{})
    for r in lt.get('rows',[]):
        s=sum(v for v in r['monthly'].values() if v is not None)
        if r.get('ytd_expense') is not None and abs(s-r['ytd_expense'])>1.0:
            ok=False; msg.append(f"{r['leader']}: sum(monthly)={round(s,1)} != YTD Exp {r['ytd_expense']}")
        if r.get('ytd_bud') is not None and r.get('ytd_expense') is not None and r.get('variance') is not None:
            if abs((r['ytd_bud']-r['ytd_expense'])-r['variance'])>1.0:
                ok=False; msg.append(f"{r['leader']}: variance {r['variance']} != YTDBud-YTDExp {round(r['ytd_bud']-r['ytd_expense'],1)}")
    return ok,(msg or ['all identity checks passed'])

if __name__=='__main__':
    path=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/outputs/Travel_Expense_Jun26_INPUT_withMeta.xlsx'
    R=convert(path); ok,msg=validate(R)
    print("META:",R['meta'])
    print("leader months:",R['leader_table']['months'],"| leader rows:",len(R['leader_table']['rows']))
    print("sojern depts:",len(R['sojern_dept']['rows']),"| entities:",R['sojern_dept']['entities'])
    print("VALIDATE:",("PASS" if ok else "FAIL"),"|",'; '.join(msg[:4]))
    json.dump(R,open((__import__('os').environ.get('STD_OUT') or '/mnt/user-data/outputs')+'/travel_STANDARDIZED.json','w'),ensure_ascii=False,indent=1)
